#
# Developer documentation: https://github.com/AdvancedPhotonSource/ComponentDB/wiki/cable-import-script-developer-notes
# User documentation: https://github.com/AdvancedPhotonSource/ComponentDB/wiki/cable-preimport-script
#
import concurrent
from concurrent.futures.thread import ThreadPoolExecutor
from datetime import datetime
import os

import urllib3

print("working directory: %s" % os.getcwd())

import argparse
import logging
import configparser
import sys
from abc import ABC, abstractmethod
import re

import openpyxl
import xlsxwriter

from CdbApiFactory import CdbApiFactory
from cdbApi import ApiException, ItemDomainCableCatalogIdListRequest, ItemDomainCableDesignIdListRequest, \
    ItemDomainMachineDesignIdListRequest, SourceIdListRequest, ConnectorTypeIdListRequest, CableCatalogItemInfo

# constants

DEFAULT_COLUMN_WIDTH = 80
DEFAULT_FONT_HEIGHT = 12

LABEL_CABLES_NAME = "Kabel name (*NOT the final APS-U, CDB, or MBA name)"
LABEL_CABLES_LAYING = "Laying"
LABEL_CABLES_VOLTAGE = "Voltage"
LABEL_CABLES_OWNER = "Owner"
LABEL_CABLES_TYPE = "Type"
LABEL_CABLES_SRC_LOCATION = "Location"
LABEL_CABLES_SRC_ANSU = "A/N/S/U"
LABEL_CABLES_SRC_ETPMC = "E/T/P/M/C"
LABEL_CABLES_SRC_ADDRESS = "Address"
LABEL_CABLES_SRC_DESCRIPTION = "Description"
LABEL_CABLES_DEST_LOCATION = "Location"
LABEL_CABLES_DEST_ANSU = "A/N/S/U"
LABEL_CABLES_DEST_ETPMC = "E/T/P/M/C"
LABEL_CABLES_DEST_ADDRESS = "Address"
LABEL_CABLES_DEST_DESCRIPTION = "Description"
LABEL_CABLES_CABLE_ID = "Cable ID"
LABEL_CABLES_END1_DEVICE = "End1 device"
LABEL_CABLES_END1_PORT = "End1 port"
LABEL_CABLES_END2_DEVICE = "End2 device"
LABEL_CABLES_END2_PORT = "End2 port"
LABEL_CABLES_IMPORT_ID = "Import cable ID"
LABEL_CABLES_FIRST_WAYPOINT = "First waypoint"
LABEL_CABLES_FINAL_WAYPOINT = "Final waypoint"
LABEL_CABLES_NOTES = "Notes"

LABEL_CABLESPECS_DESCRIPTION = "cable type description"
LABEL_CABLESPECS_MANUFACTURER = "manufacturer"
LABEL_CABLESPECS_PART_NUM = "part number"
LABEL_CABLESPECS_ALT_PART_NUM = "alt part number"
LABEL_CABLESPECS_DIAMETER = "diameter"
LABEL_CABLESPECS_WEIGHT = "weight"
LABEL_CABLESPECS_CONDUCTORS = "conductors"
LABEL_CABLESPECS_INSULATION = "insulation"
LABEL_CABLESPECS_JACKET = "jacket color"
LABEL_CABLESPECS_VOLTAGE = "voltage rating"
LABEL_CABLESPECS_FIRE_LOAD = "fire load"
LABEL_CABLESPECS_HEAT_LIMIT = "heat limit"
LABEL_CABLESPECS_BEND_RADIUS = "bend radius"
LABEL_CABLESPECS_RAD_TOLERANCE = "rad tolerance"
LABEL_CABLESPECS_LINK = "link (URL)"
LABEL_CABLESPECS_IMAGE = "image (URL)"
LABEL_CABLESPECS_TOTAL_LENGTH = "total length"
LABEL_CABLESPECS_REEL_LENGTH = "reel length"
LABEL_CABLESPECS_REEL_QUANTITY = "reel quantity"
LABEL_CABLESPECS_LEAD_TIME = "lead time"
LABEL_CABLESPECS_ORDERED = "ordered"
LABEL_CABLESPECS_RECEIVED = "received"
LABEL_CABLESPECS_CHECKLIST = "checklist"
LABEL_CABLESPECS_E1_1 = "e1-1"
LABEL_CABLESPECS_E2_1 = "e2-1"
LABEL_CABLESPECS_E1_2 = "e1-2"
LABEL_CABLESPECS_E2_2 = "e2-2"
LABEL_CABLESPECS_E1_3 = "e1-3"
LABEL_CABLESPECS_E2_3 = "e2-3"

CABLE_TYPE_NAME_KEY = "Name"
CABLE_TYPE_ALT_NAME_KEY = "Alt Name"
CABLE_TYPE_DESCRIPTION_KEY = "Description"
CABLE_TYPE_LINK_URL_KEY = "Documentation URL"
CABLE_TYPE_IMAGE_URL_KEY = "Image URL"
CABLE_TYPE_MANUFACTURER_KEY = "Manufacturer"
CABLE_TYPE_PART_NUMBER_KEY = "Part Number"
CABLE_TYPE_ALT_PART_NUMBER_KEY = "Alt Part Num"
CABLE_TYPE_DIAMETER_KEY = "Diameter"
CABLE_TYPE_WEIGHT_KEY = "Weight"
CABLE_TYPE_CONDUCTORS_KEY = "Conductors"
CABLE_TYPE_INSULATION_KEY = "Insulation"
CABLE_TYPE_JACKET_COLOR_KEY = "Jacket Color"
CABLE_TYPE_VOLTAGE_RATING_KEY = "Voltage Rating"
CABLE_TYPE_FIRE_LOAD_KEY = "Fire Load"
CABLE_TYPE_HEAT_LIMIT_KEY = "Heat Limit"
CABLE_TYPE_BEND_RADIUS_KEY = "Bend Radius"
CABLE_TYPE_RAD_TOLERANCE_KEY = "Rad Tolerance"
CABLE_TYPE_TOTAL_LENGTH_KEY = "Total Length"
CABLE_TYPE_REEL_LENGTH_KEY = "Reel Length"
CABLE_TYPE_REEL_QTY_KEY = "Reel Quantity"
CABLE_TYPE_LEAD_TIME_KEY = "Lead Time"
CABLE_TYPE_ORDERED_KEY = "ordered"
CABLE_TYPE_RECEIVED_KEY = "received"
CABLE_TYPE_CHECKLIST_KEY = "checklist"
CABLE_TYPE_COLUMN_Y_KEY = "column Y"
CABLE_TYPE_COLUMN_Z_KEY = "column Z"
CABLE_TYPE_E1_1_KEY = "e1-1"
CABLE_TYPE_E2_1_KEY = "e2-1"
CABLE_TYPE_E1_2_KEY = "e1-2"
CABLE_TYPE_E2_2_KEY = "e2-2"
CABLE_TYPE_E1_3_KEY = "e1-3"
CABLE_TYPE_E2_3_KEY = "e2-3"

CABLE_INVENTORY_NAME_KEY = "name"

CABLE_DESIGN_NAME_KEY = "name"
CABLE_DESIGN_LAYING_KEY = "laying"
CABLE_DESIGN_VOLTAGE_KEY = "voltage"
CABLE_DESIGN_OWNER_KEY = "owner"
CABLE_DESIGN_TYPE_KEY = "type"
CABLE_DESIGN_SRC_LOCATION_KEY = "srcLocation"
CABLE_DESIGN_SRC_ANS_KEY = "srcANS"
CABLE_DESIGN_SRC_ETPM_KEY = "srcETPM"
CABLE_DESIGN_SRC_ADDRESS_KEY = "srcAddress"
CABLE_DESIGN_SRC_DESCRIPTION_KEY = "srcDescription"
CABLE_DESIGN_DEST_LOCATION_KEY = "destLocation"
CABLE_DESIGN_DEST_ANS_KEY = "destANS"
CABLE_DESIGN_DEST_ETPM_KEY = "destETPM"
CABLE_DESIGN_DEST_ADDRESS_KEY = "destAddress"
CABLE_DESIGN_DEST_DESCRIPTION_KEY = "destDescription"
CABLE_DESIGN_CABLE_ID_KEY = "cableId"
CABLE_DESIGN_END1_DEVICE_NAME_KEY = "end1Device"
CABLE_DESIGN_END1_PORT_NAME_KEY = "end1Port"
CABLE_DESIGN_END2_DEVICE_NAME_KEY = "end2Device"
CABLE_DESIGN_END2_PORT_NAME_KEY = "end2Port"
CABLE_DESIGN_IMPORT_ID_KEY = "importId"
CABLE_DESIGN_VIA_ROUTE_KEY = "via"
CABLE_DESIGN_WAYPOINT_ROUTE_KEY = "waypoint"
CABLE_DESIGN_NOTES_KEY = "notes"


class ConnectedMenuManager:

    def __init__(self, workbook):
        self.name_dict = {}
        self.initialize(workbook)

    def add_name(self, name, values):
        self.name_dict[name] = values

    def initialize(self, workbook):

        error_messages = []
        warning_messages = []
        for name in workbook.defined_names.definedName:
            range_name = name.name
            if range_name == "_Print_Area" or range_name == "_Rack_data_201901141042_usmani":
                continue
            expression = name.value
            if '!' not in expression:
                error_messages.append("unexpected expression format (should include '!'), range: %s expression: %s" % (range_name, expression))
                continue
            (sheet_name, ref) = expression.split('!')
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                error_messages.append("invalid sheet name in expression, range: %s expression: %s sheet: %s" % (range_name, expression, sheet_name))
                continue
            if ':' in ref:
                (first_cell, last_cell) = ref.split(':')
            else:
                first_cell = ref
                last_cell = ref
            (first_cell_row, first_cell_col) = xl_cell_to_rowcol(first_cell)
            (last_cell_row, last_cell_col) = xl_cell_to_rowcol(last_cell)
            if first_cell_row < 0 or first_cell_col < 0 or last_cell_row < first_cell_row or last_cell_col < first_cell_col or last_cell_row > sheet.max_row or last_cell_col > sheet.max_column:
                error_messages.append("invalid indexes for range, sheet: %s name: %s ref: %s first_cell_row: %d first_cell_col: %d last_cell_row: %d last_cell_col: %d numrows: %d numcols: %d" % (sheet_name, range_name, ref, first_cell_row, first_cell_col, last_cell_row, last_cell_col, sheet.max_row, sheet.max_column))
                continue
            values = []
            has_error = False
            for row_ind in range(first_cell_row, last_cell_row + 1):
                for col_ind in range(first_cell_col, last_cell_col + 1):
                    try:
                        cell = sheet.cell(row_ind, col_ind)
                        cell_value = cell.value
                        if cell_value is None or cell_value == "":
                            # ignore blank value in named range
                            break
                        else:
                            if cell_value != range_name:
                                # don't add range name as a a value, this is treated inconsistently in the workbooks
                                values.append(cell_value)
                    except IndexError:
                        error_messages.append("cell index error, sheet: %s range: %s row: %d col: %d" % (sheet_name, name.name, row_ind+1, col_ind+1))
                        has_error = True
                        break
                if has_error:
                    break
            if has_error:
                continue
            if range_name in values:
                error_messages.append("sheet: %s range: %s ref: %s includes range name in values" % (sheet_name, range_name, ref))
                continue
            self.add_name(range_name, values)
            # else:
            #     if range_name != "_Print_Area":
            #         warning_messages.append("ignoring range not at global scope (=-1): %s scope: %s expression: %s" % (range_name, name.scope, name.formula_text))
        if len(warning_messages) > 0:
            print()
            print("EXCEL NAMED RANGE WARNINGS ====================")
            print()
            for message in warning_messages:
                print(message)
        if len(error_messages) > 0:
            print()
            print("EXCEL NAMED RANGE ERRORS ====================")
            print()
            for message in error_messages:
                print(message)
            fatal_error("Error(s) parsing excel named ranges, see console for details. Exiting.")

    def has_name(self, range_name):
        return range_name in self.name_dict

    def value_is_valid_for_name(self, parent_value, child_value):
        child_value_list = self.name_dict.get(parent_value)
        if child_value_list is not None:
            return child_value in child_value_list
        else:
            return False

    def num_values_for_name(self, range_name):
        if range_name in self.name_dict:
            return len(self.name_dict[range_name])
        return 0

    def values_for_name(self, range_name):
        if range_name in self.name_dict:
            return self.name_dict[range_name]
        return []


class IdManager():

    def __init__(self):
        self.name_id_dict = {}

    def set_dict(self, dict):
        self.name_id_dict = dict

    def update(self, dict):
        self.name_id_dict.update(dict)

    def set_id_for_name(self, name, id):
        self.name_id_dict[name] = id

    def get_id_for_name(self, name):
        if name in self.name_id_dict:
            return self.name_id_dict[name]
        else:
            return None


class RackManager():

    def __init__(self):
        self.rack_dict = {}

    def add_endpoint_id_for_rack(self, rack_name, endpoint_name, endpoint_id):
        if not rack_name in self.rack_dict:
            self.rack_dict[rack_name] = {}
        rack_contents = self.rack_dict[rack_name]
        rack_contents[endpoint_name] = endpoint_id

    def get_endpoint_id_for_rack(self, rack_name, endpoint_name):
        if rack_name in self.rack_dict:
            rack_items = self.rack_dict[rack_name]
            if endpoint_name in rack_items:
                return rack_items[endpoint_name]
        return None


class ItemInfoManager():

    def __init__(self, api):
        self.api = api
        self.source_id_manager = IdManager()
        self.output_objects_source = []
        self.existing_sources = set()
        self.new_sources = set()
        self.existing_connector_types = set()
        self.new_connector_types = set()
        self.output_connector_types = []
        self.output_cable_type_connectors = []
        self.cable_type_names = []
        self.existing_cable_types = []
        self.new_cable_types = []
        self.undefined_cable_types = []
        self.cable_type_info = {}
        self.technical_system = None
        self.cable_types_specified_for_technical_system = None
        self.cable_types_defined_for_technical_system = []
        self.output_objects_cable_type_compare = []
        self.cable_design_names = []
        self.existing_cable_designs = []
        self.new_cable_designs = []
        self.id_manager_cable_design = IdManager()
        self.machine_info = {}
        self.rack_manager = RackManager()
        self.missing_endpoints = set()
        self.nonunique_endpoints = set()
        self.from_port_values = []
        self.to_port_values = []
        self.catalog_ports = {}
        self.output_objects_cable_design_compare = []
        self.output_objects_cable_inventory = []
        self.name_manager = None

    def initialize(self, input_book):
        self.name_manager = ConnectedMenuManager(input_book)

    def set_technical_system(self, technical_system):
        self.technical_system = technical_system

    def get_technical_system(self):
        return self.technical_system

    def set_cable_types_specified_for_technical_system(self, cable_type_list):
        self.cable_types_specified_for_technical_system = cable_type_list

    def get_cable_types_specified_for_technical_system(self):
        return self.cable_types_specified_for_technical_system

    def add_cable_type_defined_for_technical_system(self, cable_type):
        self.cable_types_defined_for_technical_system.append(cable_type)

    def get_cable_types_defined_for_technical_system(self):
        return self.cable_types_defined_for_technical_system

    def get_connector_type_id_list(self, connector_type_name_list):

        connector_type_name_list = list(connector_type_name_list)

        print("fetching %d connector type id's" % (len(connector_type_name_list)))

        try:
            request_obj = ConnectorTypeIdListRequest(name_list=connector_type_name_list)
            id_list = self.api.getConnectorTypesApi().get_connector_type_id_list(
                connector_type_id_list_request=request_obj)
        except ApiException as ex:
            fatal_error("unknown api exception getting list of connector type ids")

        # list sizes should match
        if len(connector_type_name_list) != len(id_list):
            fatal_error("api list size mismatch getting list of connector type ids")

        print("fetched %d connector type id's" % (len(id_list)))

        return id_list

    def initialize_connector_types(self, connector_type_names):
        id_list = self.get_connector_type_id_list(connector_type_names)
        for connector_type_name, connector_type_id in zip(connector_type_names, id_list):
            if connector_type_id != 0:
                self.existing_connector_types.add(connector_type_name)
            else:
                if connector_type_name not in self.new_connector_types:
                    self.new_connector_types.add(connector_type_name)
                    self.output_connector_types.append(ConnectorTypeOutputObject(connector_type_name))

    def add_output_cable_type_connector(self, output_cable_type_connector):
        self.output_cable_type_connectors.append(output_cable_type_connector)

    def get_cable_type_info_list(self, cable_type_name_list):

        cable_type_name_list = list(cable_type_name_list)

        print("fetching %d cable type info objects" % (len(cable_type_name_list)))

        try:
            request_obj = ItemDomainCableCatalogIdListRequest(name_list=cable_type_name_list)
            info_list = self.api.getCableImportApi().get_cable_catalog_info_list(
                item_domain_cable_catalog_id_list_request=request_obj)
        except ApiException as ex:
            fatal_error("unknown api exception getting list of cable type info objects")

        # list sizes should match
        if len(cable_type_name_list) != len(info_list):
            fatal_error("api list size mismatch getting list of cable type info objects")

        print("fetched %d cable type info objects" % (len(info_list)))

        return info_list

    def add_cable_type_info(self, cable_type_name, cable_type_info):
        self.cable_type_info[cable_type_name] = cable_type_info

    def load_cable_type_info(self, cable_type_names):
        info_list = self.get_cable_type_info_list(cable_type_names)
        for cable_type_name, cable_type_info in zip(cable_type_names, info_list):
            if cable_type_info is None:
                # cable type doesn't exist in cdb, create info object with id=0
                cable_type_info = CableCatalogItemInfo()
                cable_type_info.name = cable_type_name
                cable_type_info.id = 0
                cable_type_info.connector_names = []
            self.add_cable_type_info(cable_type_name, cable_type_info)

    def get_cable_type_info(self, cable_type_name):
        if cable_type_name not in self.cable_type_info:
            return None
        else:
            return self.cable_type_info[cable_type_name]

    def get_cable_type_id(self, cable_type_name):
        cable_type_info = self.get_cable_type_info(cable_type_name)
        if cable_type_info is None:
            return 0
        else:
            return cable_type_info.id

    def handle_cable_type_connector(self, cable_type_name, connector_name, cable_end, connector_type):
        if cable_type_name in self.cable_type_info.keys():
            cable_type_info = self.cable_type_info[cable_type_name]
            if cable_type_info.id == 0:
                # only handle connector for new cable catalog items, ignore for existing ones
                cable_type_info.connector_names.append(connector_name)
                output_object = CableTypeConnectorOutputObject(cable_type_name, connector_name, cable_end, connector_type)
                self.add_output_cable_type_connector(output_object)

    def add_machine_info(self, machine_item_name, machine_info):
        if machine_item_name not in self.machine_info.keys():
            self.machine_info[machine_item_name] = machine_info

    def get_machine_info(self, machine_item_name):
        if machine_item_name in self.machine_info.keys():
            return self.machine_info[machine_item_name]
        else:
            return None

    def add_catalog_port(self, catalog_name, port_name):
        if catalog_name not in self.catalog_ports.keys():
            self.catalog_ports[catalog_name] = []
        port_names = self.catalog_ports[catalog_name]
        if port_name not in port_names:
            port_names.append(port_name)

    def get_catalog_ports(self, catalog_name):
        if catalog_name in self.catalog_ports.keys():
            return self.catalog_ports[catalog_name]
        else:
            return []

    def get_output_objects_cable_design_compare(self):
        return self.output_objects_cable_design_compare

    def get_output_objects_cable_inventory(self):
        return self.output_objects_cable_inventory


class InputHandler(ABC):

    def __init__(self, column_key, info_mgr=None):
        self.column_key = column_key
        self.info_manager = info_mgr

    # initializes handler, subclasses override to customize
    def initialize(self, api, sheet, first_row, last_row):
        pass

    # Invokes handler.
    @abstractmethod
    def handle_input(self, input_dict):
        pass


class ConnectedMenuHandler(InputHandler):

    def __init__(self, info_manager, column_key, parent_key):
        super().__init__(column_key, info_manager)
        self.parent_key = parent_key

    def handle_input(self, input_dict):
        parent_value = input_dict[self.parent_key]
        cell_value = input_dict[self.column_key]
        if not self.info_manager.name_manager.has_name(parent_value):
            return False, "name manager has no menu range for: %s column: %s parent column: %s" % (parent_value, self.column_key, self.parent_key)
        has_child = self.info_manager.name_manager.value_is_valid_for_name(parent_value, cell_value)
        valid_string = ""
        if not has_child:
            valid_string = "range for parent name %s does not include child name %s" % (parent_value, cell_value)
        return has_child, valid_string


class NamedRangeHandler(InputHandler):

    def __init__(self, info_manager, column_key, range_name):
        super().__init__(column_key, info_manager)
        self.range_name = range_name

    def handle_input(self, input_dict):
        if not self.info_manager.name_manager.has_name(self.range_name):
            return False, "name manager has no named range for: %s" % self.range_name
        cell_value = input_dict[self.column_key]
        has_child = self.info_manager.name_manager.value_is_valid_for_name(self.range_name, cell_value)
        valid_string = ""
        if not has_child:
            valid_string = "named range %s does not include value %s" % (self.range_name, cell_value)
        return has_child, valid_string


class TechnicalSystemCableTypeValidationHandler(InputHandler):

    def __init__(self, column_key, info_manager):
        super().__init__(column_key, info_manager)

    def handle_input(self, input_dict):
        is_valid = True
        valid_string = ""
        cell_value = input_dict[self.column_key]
        if cell_value not in self.info_manager.get_cable_types_specified_for_technical_system():
            technical_system = self.info_manager.get_technical_system()
            is_valid = False
            valid_string = "cable type: %s from CableSpecs tab not specified for technical system: %s in CableTypes tab" % (cell_value, technical_system)
        else:
            self.info_manager.add_cable_type_defined_for_technical_system(cell_value)
        return is_valid, valid_string


class UniqueNameHandler(InputHandler):

    def __init__(self, column_key, item_names):
        super().__init__(column_key)
        self.item_names = item_names

    def handle_input(self, input_dict):

        item_name = input_dict[self.column_key]

        if item_name in self.item_names:
            # flag duplicate item name
            error_msg = "duplicate value: %s for column: %s" % (item_name, self.column_key)
            logging.error(error_msg)
            return False, error_msg
        else:
            self.item_names.append(item_name)

        return True, ""


class DeviceAddressHandler(InputHandler):

    def __init__(self, info_manager, column_key, location_key, etpm_key):
        super().__init__(column_key, info_manager)
        self.location_key = location_key
        self.etpm_key = etpm_key

    def handle_input(self, input_dict):

        location_value = input_dict[self.location_key]
        etpm_value = input_dict[self.etpm_key]
        cell_value = input_dict[self.column_key]

        range_name = ""
        if location_value == "SR_T":
            range_name = "Snn" + etpm_value[3:]
        elif "PS-" in etpm_value:
            range_name = "_PS_CAB_SLOT_"
        else:
            range_name = "_RACK_AREA_"

        if not self.info_manager.name_manager.has_name(range_name):
            return False, "name manager has no named address range for: %s" % range_name

        has_child = self.info_manager.name_manager.value_is_valid_for_name(range_name, cell_value)
        valid_string = ""
        if not has_child:
            valid_string = "named address range %s does not include value %s" % (range_name, cell_value)
        return has_child, valid_string


class EndpointHandler(InputHandler):

    def __init__(self, column_key, rack_key, hierarchy_name, info_mgr, api, column_index_item_name, column_index_rack_name, description):
        super().__init__(column_key, info_mgr)
        self.rack_key = rack_key
        self.hierarchy_name = hierarchy_name
        self.api = api
        self.column_index_item_name = column_index_item_name
        self.column_index_rack_name = column_index_rack_name
        self.description = description

    def call_api(self, api, item_names_batch, rack_names_batch):
        request_obj = ItemDomainMachineDesignIdListRequest(item_names=item_names_batch,
                                                           rack_names=rack_names_batch,
                                                           root_name=self.hierarchy_name)
        info_list = self.api.getCableImportApi().get_machine_info_list(
            item_domain_machine_design_id_list_request=request_obj)
        return info_list

    def initialize(self, api, sheet, first_row, last_row):

        # create map of item name to list of rack names (in case the same item name is used in more than one rack)
        rack_items_dict = {}
        for row_ind in range(first_row, last_row+1):
            # increment column index values since cell() is 1-based
            item_name = sheet.cell(row_ind, self.column_index_item_name+1).value
            rack_name = sheet.cell(row_ind, self.column_index_rack_name+1).value
            if (item_name is not None and item_name != "") and (rack_name is not None and rack_name != ""):
                if rack_name not in rack_items_dict:
                    rack_items_dict[rack_name] = []
                rack_items = rack_items_dict[rack_name]
                if item_name not in rack_items:
                    rack_items.append(item_name)

        # create parallel lists of item and rack names for calling api
        rack_names = []
        item_names = []
        for rack_name in rack_items_dict:
            rack_items = rack_items_dict[rack_name]
            for item_name in rack_items:
                rack_names.append(rack_name)
                item_names.append(item_name)

        if len(item_names) != len(rack_names):
            fatal_error("error preparing lists for machine item id API")

        print("fetching %d machine item id's for %s" % (len(item_names), self.description))

        preapi_time = datetime.now()

        max_workers = 5  # using value larger than 5 causes RemoteDisconnected exception
        batch_size = 250
        num_iterations = len(item_names) // batch_size
        if len(item_names) % batch_size != 0:
            num_iterations = num_iterations + 1

        with ThreadPoolExecutor(max_workers=max_workers) as executor:

            futures = {}
            iteration_num = 1
            start_index = 0
            end_index = batch_size
            result_id_count = 0
            while iteration_num <= num_iterations:

                if end_index > len(item_names):
                    end_index = len(item_names)

                item_names_batch = item_names[start_index:end_index]
                rack_names_batch = rack_names[start_index:end_index]

                future = executor.submit(self.call_api, api, item_names_batch, rack_names_batch)
                futures[future] = (iteration_num, item_names_batch, rack_names_batch, start_index, end_index)

                iteration_num = iteration_num + 1
                start_index = start_index + batch_size
                end_index = end_index + batch_size

            for future in concurrent.futures.as_completed(futures):
                try:
                    info_list = future.result()

                except ApiException as ex:
                    fatal_error("unknown api exception getting list of machine item ids")

                except Exception as exc:
                    fatal_error("exception calling machine item api: %s" % exc)

                else:
                    (iteration_num, item_names_batch, rack_names_batch, start_index, end_index) = futures[future]
                    # list sizes should match
                    if len(info_list) != len(item_names_batch):
                        fatal_error("api result list size mismatch getting list of machine item info")
                    # iterate 3 lists to process api result
                    for (item_name, rack_name, machine_info) in zip(item_names_batch, rack_names_batch, info_list):
                        id = machine_info.id
                        self.info_manager.rack_manager.add_endpoint_id_for_rack(rack_name, item_name, id)
                        self.info_manager.add_machine_info(item_name, machine_info)
                    result_id_count = result_id_count + len(info_list)
                    print("fetched %d machine item id's for %s" % (result_id_count, self.description))

        end_time = datetime.now()
        print("machine item id api duration: %d sec." % (end_time-preapi_time).total_seconds())

    def handle_input(self, input_dict):

        endpoint_name = input_dict[self.column_key]
        rack_name = input_dict[self.rack_key]

        is_valid = True
        valid_string = ""

        id = self.info_manager.rack_manager.get_endpoint_id_for_rack(rack_name, endpoint_name)
        if id == 0:
            is_valid = False
            valid_string = "no endpoint item found in CDB with name: %s rack: %s in hierarchy: %s" % (endpoint_name, rack_name, self.hierarchy_name)
            logging.error(valid_string)
            self.info_manager.missing_endpoints.add("rack: %s device: %s" % (rack_name, endpoint_name))
        elif id == -1:
            is_valid = False
            valid_string = "duplicate endpoint items found in CDB with name: %s rack: %s in hierarchy: %s" % (endpoint_name, rack_name, self.hierarchy_name)
            logging.error(valid_string)
            self.info_manager.nonunique_endpoints.add(rack_name + " + " + endpoint_name)
        else:
            logging.debug("found machine design item in CDB with name: %s, id: %s" % (endpoint_name, id))

        return is_valid, valid_string


class CableTypeExistenceHandler(InputHandler):

    def __init__(self, info_manager, column_key):
        super().__init__(column_key, info_manager)

    def initialize(self, api, sheet, first_row, last_row):
        pass

    def handle_input(self, input_dict):
        cable_type_name = input_dict[self.column_key]
        cable_type_id = self.info_manager.get_cable_type_id(cable_type_name)
        if cable_type_id is None:
            return False, "unexpected error in id map for cable type existence check"
        if cable_type_id != 0:
            # cable type already exists
            self.info_manager.existing_cable_types.append(cable_type_name)
            return True, ""
        else:
            self.info_manager.new_cable_types.append(cable_type_name)
            return True, ""


class CableTypeValidForTechnicalSystemHandler(InputHandler):

    def __init__(self, info_manager, column_key):
        super().__init__(column_key, info_manager)

    def initialize(self, api, sheet, first_row, last_row):
        pass

    def handle_input(self, input_dict):

        cable_type_name = input_dict[self.column_key]

        if cable_type_name not in self.info_manager.get_cable_types_specified_for_technical_system():
            return False, "cable type: %s not specified in CableTypes column for technical system: %s" % (cable_type_name, self.info_manager.get_technical_system())
        else:
            return True, ""


class CableTypeConnectorHandler(InputHandler):

    def __init__(self, column_key, cable_end):
        super().__init__(column_key)
        self.cable_end = cable_end

    def initialize(self, api, sheet, first_row, last_row):
        pass

    def handle_input(self, input_dict):

        connector_type = input_dict[self.column_key]

        if connector_type is not None and connector_type != "":
            cable_type_name = input_dict[CABLE_TYPE_NAME_KEY]
            connector_name = self.column_key
            self.info_manager.handle_cable_type_connector(cable_type_name, connector_name, self.cable_end, connector_type)

        return True, ""


class CableDesignExistenceHandler(InputHandler):

    def __init__(self, column_key, info_mgr, column_index_cable_id, column_index_import_id):
        super().__init__(column_key, info_mgr)
        self.id_mgr = self.info_manager.id_manager_cable_design
        self.column_index_cable_id = column_index_cable_id
        self.column_index_import_id = column_index_import_id

    def call_api(self, api, cable_design_names_batch):

        request_obj = ItemDomainCableDesignIdListRequest(name_list=cable_design_names_batch)
        id_list = api.getCableDesignItemApi().get_cable_design_id_list(item_domain_cable_design_id_list_request=request_obj)
        return id_list

    def initialize(self, api, sheet, first_row, last_row):

        cable_design_names = []
        for row_ind in range(first_row, last_row+1):
            cable_id = sheet.cell(row_ind, self.column_index_cable_id+1).value
            import_id = sheet.cell(row_ind, self.column_index_import_id+1).value
            cable_design_names.append(CableDesignOutputObject.get_name_cls(None, cable_id, import_id))
        print("fetching %d cable design id's" % len(cable_design_names))

        preapi_time = datetime.now()
        max_workers = 5  # setting to value greater than 5 leads to RemoteDisconnected exception
        batch_size = 250
        num_iterations = len(cable_design_names) // batch_size
        if len(cable_design_names) % batch_size != 0:
            num_iterations = num_iterations + 1

        with ThreadPoolExecutor(max_workers=max_workers) as executor:

            futures = {}
            iteration_num = 1
            start_index = 0
            end_index = batch_size
            result_id_count = 0

            while iteration_num <= num_iterations:

                if end_index > len(cable_design_names):
                    end_index = len(cable_design_names)

                cable_design_names_batch = cable_design_names[start_index:end_index]

                future = executor.submit(self.call_api, api, cable_design_names_batch)
                futures[future] = (iteration_num, cable_design_names_batch, start_index, end_index)

                iteration_num = iteration_num + 1
                start_index = start_index + batch_size
                end_index = end_index + batch_size

            for future in concurrent.futures.as_completed(futures):
                try:
                    id_list = future.result()
                except ApiException as ex:
                    fatal_error("unknown api exception getting list of cable design ids")
                except Exception as exc:
                    fatal_error("exception calling cable design api: %s" % exc)
                else:
                    (iteration_num, cable_design_names_batch, start_index, end_index) = futures[future]
                    # list sizes should match
                    if len(cable_design_names_batch) != len(id_list):
                        fatal_error("api result list size mismatch getting list of cable design ids")
                    result_dict = dict(zip(cable_design_names_batch, id_list))
                    self.id_mgr.update(result_dict)
                    result_id_count = result_id_count + len(id_list)
                    print("fetched %d cable design ids" % result_id_count)

        end_time = datetime.now()
        print("cable design id api duration: %d sec." % (end_time-preapi_time).total_seconds())

    def handle_input(self, input_dict):
        cable_design_name = CableDesignOutputObject.get_name_cls(input_dict)
        # check to see if cable design exists in CDB
        cable_design_id = self.id_mgr.get_id_for_name(cable_design_name)
        if cable_design_id is None:
            return False, "unexpected error with missing entry in cable design id map"
        if cable_design_id != 0:
            # cable design already exists
            self.info_manager.existing_cable_designs.append(cable_design_name)
        else:
            self.info_manager.new_cable_designs.append(cable_design_name)
        return True, ""


class DevicePortHandler(InputHandler):

    def __init__(self, column_key, endpoint_key, info_mgr, ignore_port_values, values):
        super().__init__(column_key, info_mgr)
        self.endpoint_key = endpoint_key
        self.ignore_port_values = ignore_port_values
        self.values = values

    def initialize(self, api, sheet, first_row, last_row):
        pass

    def handle_input(self, input_dict):

        cell_value = input_dict[self.column_key]
        endpoint_name = input_dict[self.endpoint_key]

        if self.ignore_port_values:
            if cell_value is not None and cell_value != "":
                self.values.append(cell_value)

        # save port information to info manager data structure for use in generating CDB catalog ports tab in output workbook
        machine_info = self.info_manager.get_machine_info(endpoint_name)
        if machine_info is not None:
            catalog_name = machine_info.catalog_name
            if catalog_name is not None and catalog_name != "":
                self.info_manager.add_catalog_port(catalog_name, cell_value)

        return True, ""


class SourceHandler(InputHandler):

    def __init__(self, info_mgr, column_key, column_index, api):
        super().__init__(column_key, info_mgr)
        self.column_index = column_index
        self.api = api

    @staticmethod
    def get_source_id_list(api, source_name_list, description):

        source_name_list = list(source_name_list)

        print("fetching %d source id's for %s" % (len(source_name_list), description))

        try:
            request_obj = SourceIdListRequest(name_list=source_name_list)
            id_list = api.getSourceApi().get_source_id_list(source_id_list_request=request_obj)
        except ApiException as ex:
            fatal_error("unknown api exception getting list of source ids for %s" % description)

        # list sizes should match
        if len(source_name_list) != len(id_list):
            fatal_error("api list size mismatch getting list of source ids for %s" % description)

        print("fetched %d source id's for %s" % (len(id_list), description))

        return id_list

    def initialize(self, api, sheet, first_row, last_row):

        source_names = set()  # use set to eliminate duplicates

        for row_ind in range(first_row, last_row+1):
            # cell() is 1-based so increment index for access
            val = sheet.cell(row_ind, self.column_index+1).value
            if val is not None and val != "":
                source_names.add(val)

        id_list = SourceHandler.get_source_id_list(api, source_names, "manufacturer lookup")

        self.info_manager.source_id_manager.set_dict(dict(zip(source_names, id_list)))

    def handle_input(self, input_dict):

        source_name = input_dict[self.column_key]

        if len(source_name) == 0:
            # no manufacturer specified
            return True, ""

        cached_id = self.info_manager.source_id_manager.get_id_for_name(source_name)
        if cached_id != 0:
            self.info_manager.existing_sources.add(source_name)
            logging.debug("found source with name: %s, id: %s" % (source_name, cached_id))
        else:
            if source_name not in self.info_manager.new_sources:
                self.info_manager.new_sources.add(source_name)
                self.info_manager.output_objects_source.append(SourceOutputObject(None, input_dict))
                logging.debug("adding new source: %s" % source_name)
            else:
                logging.debug("already added new source: %s" % source_name)

        return True, ""


class InputColumnModel:

    def __init__(self, key=None, label=None, required=False):
        self.key = key
        self.required = required
        self.label = label


class OutputColumnModel:

    def __init__(self, method, label=""):
        self.method = method
        self.label = label


class OutputObject(ABC):

    def __init__(self, helper, input_dict):
        self.helper = helper
        self.input_dict = input_dict

    @classmethod
    @abstractmethod
    def get_output_columns(cls):
        pass

    def empty_column(self):
        return ""


class ConnectorTypeOutputObject(OutputObject):

    def __init__(self, name):
        self.name = name

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="empty_column", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_name", label="Name"),
            OutputColumnModel(method="empty_column", label="Description"),
        ]
        return column_list

    def get_name(self):
        return self.name


class CableTypeConnectorOutputObject(OutputObject):

    def __init__(self, catalog_item, connector_name, cable_end, connector_type):
        self.catalog_item = catalog_item
        self.connector_name = connector_name
        self.cable_end = cable_end
        self.connector_type = connector_type

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="empty_column", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_catalog_item", label="Cable Catalog Item"),
            OutputColumnModel(method="get_connector_name", label="Connector Name"),
            OutputColumnModel(method="get_cable_end", label="Cable End"),
            OutputColumnModel(method="empty_column", label="Description"),
            OutputColumnModel(method="get_connector_type", label="Connector Type"),
        ]
        return column_list

    def get_catalog_item(self):
        return "#" + self.catalog_item

    def get_connector_name(self):
        return self.connector_name

    def get_cable_end(self):
        return self.cable_end

    def get_connector_type(self):
        return "#" + self.connector_type


class CatalogPortOutputObject(OutputObject):

    def __init__(self, catalog_item_name, port_name):
        self.catalog_item_name = catalog_item_name
        self.port_name = port_name

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="empty_column", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_catalog_item_name", label="Catalog Item"),
            OutputColumnModel(method="get_port_name", label="Port Name"),
            OutputColumnModel(method="empty_column", label="Description"),
            OutputColumnModel(method="empty_column", label="Connector Type"),
        ]
        return column_list

    def get_catalog_item_name(self):
        return "#" + self.catalog_item_name

    def get_port_name(self):
        return self.port_name


class SourceOutputObject(OutputObject):

    def __init__(self, helper, input_dict):
        super().__init__(helper, input_dict)
        self.description = ""
        self.contact_info = ""
        self.url = ""

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="get_existing_item_id", label="Existing Item ID"),
            OutputColumnModel(method="get_delete_existing_item", label="Delete Existing Item"),
            OutputColumnModel(method="get_name", label="Name"),
            OutputColumnModel(method="get_description", label="Description"),
            OutputColumnModel(method="get_contact_info", label="Contact Info"),
            OutputColumnModel(method="get_url", label="URL"),
        ]
        return column_list

    def get_existing_item_id(self):
        return ""

    def get_delete_existing_item(self):
        return ""

    def get_name(self):
        return self.input_dict[CABLE_TYPE_MANUFACTURER_KEY]

    def get_description(self):
        return self.description

    def get_contact_info(self):
        return self.contact_info

    def get_url(self):
        return self.url


class CableTypeOutputObject(OutputObject):

    def __init__(self, helper, input_dict, existing_item_id=None):
        super().__init__(helper, input_dict)
        self.existing_item_id = existing_item_id

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="get_existing_item_id", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_name", label=CABLE_TYPE_NAME_KEY),
            OutputColumnModel(method="get_alt_name", label=CABLE_TYPE_ALT_NAME_KEY),
            OutputColumnModel(method="get_description", label=CABLE_TYPE_DESCRIPTION_KEY),
            OutputColumnModel(method="get_link_url", label=CABLE_TYPE_LINK_URL_KEY),
            OutputColumnModel(method="get_image_url", label=CABLE_TYPE_IMAGE_URL_KEY),
            OutputColumnModel(method="get_manufacturer_id", label=CABLE_TYPE_MANUFACTURER_KEY),
            OutputColumnModel(method="get_part_number", label=CABLE_TYPE_PART_NUMBER_KEY),
            OutputColumnModel(method="get_alt_part_number", label=CABLE_TYPE_ALT_PART_NUMBER_KEY),
            OutputColumnModel(method="get_diameter", label=CABLE_TYPE_DIAMETER_KEY),
            OutputColumnModel(method="get_weight", label=CABLE_TYPE_WEIGHT_KEY),
            OutputColumnModel(method="get_conductors", label=CABLE_TYPE_CONDUCTORS_KEY),
            OutputColumnModel(method="get_insulation", label=CABLE_TYPE_INSULATION_KEY),
            OutputColumnModel(method="get_jacket_color", label=CABLE_TYPE_JACKET_COLOR_KEY),
            OutputColumnModel(method="get_voltage_rating", label=CABLE_TYPE_VOLTAGE_RATING_KEY),
            OutputColumnModel(method="get_fire_load", label=CABLE_TYPE_FIRE_LOAD_KEY),
            OutputColumnModel(method="get_heat_limit", label=CABLE_TYPE_HEAT_LIMIT_KEY),
            OutputColumnModel(method="get_bend_radius", label=CABLE_TYPE_BEND_RADIUS_KEY),
            OutputColumnModel(method="get_rad_tolerance", label=CABLE_TYPE_RAD_TOLERANCE_KEY),
            OutputColumnModel(method="get_total_length", label=CABLE_TYPE_TOTAL_LENGTH_KEY),
            OutputColumnModel(method="get_reel_length", label=CABLE_TYPE_REEL_LENGTH_KEY),
            OutputColumnModel(method="get_reel_qty", label=CABLE_TYPE_REEL_QTY_KEY),
            OutputColumnModel(method="get_lead_time", label=CABLE_TYPE_LEAD_TIME_KEY),
            OutputColumnModel(method="get_procurement_status", label="Procurement Status"),
            OutputColumnModel(method="get_project_id", label="Project"),
            OutputColumnModel(method="get_tech_system_id", label="Technical System"),
            OutputColumnModel(method="get_owner_user_id", label="Owner User"),
            OutputColumnModel(method="get_owner_group_id", label="Owner Group"),
        ]
        return column_list

    def get_existing_item_id(self):
        return self.existing_item_id

    def get_name(self):
        return self.input_dict[CABLE_TYPE_NAME_KEY]

    def get_alt_name(self):
        return None

    def get_description(self):
        return self.input_dict[CABLE_TYPE_DESCRIPTION_KEY]

    def get_link_url(self):
        return self.input_dict[CABLE_TYPE_LINK_URL_KEY]

    def get_image_url(self):
        return self.input_dict[CABLE_TYPE_IMAGE_URL_KEY]

    def get_manufacturer_id(self):
        source_name = self.input_dict[CABLE_TYPE_MANUFACTURER_KEY]
        if source_name is not None and len(source_name) > 0:
            return "#" + source_name
        else:
            return None

    def get_part_number(self):
        return self.input_dict[CABLE_TYPE_PART_NUMBER_KEY]

    def get_alt_part_number(self):
        return self.input_dict[CABLE_TYPE_ALT_PART_NUMBER_KEY]

    def get_diameter(self):
        return self.input_dict[CABLE_TYPE_DIAMETER_KEY]

    def get_weight(self):
        return self.input_dict[CABLE_TYPE_WEIGHT_KEY]

    def get_conductors(self):
        return self.input_dict[CABLE_TYPE_CONDUCTORS_KEY]

    def get_insulation(self):
        return self.input_dict[CABLE_TYPE_INSULATION_KEY]

    def get_jacket_color(self):
        return self.input_dict[CABLE_TYPE_JACKET_COLOR_KEY]

    def get_voltage_rating(self):
        return self.input_dict[CABLE_TYPE_VOLTAGE_RATING_KEY]

    def get_fire_load(self):
        return self.input_dict[CABLE_TYPE_FIRE_LOAD_KEY]

    def get_heat_limit(self):
        return self.input_dict[CABLE_TYPE_HEAT_LIMIT_KEY]

    def get_bend_radius(self):
        return self.input_dict[CABLE_TYPE_BEND_RADIUS_KEY]

    def get_rad_tolerance(self):
        return self.input_dict[CABLE_TYPE_RAD_TOLERANCE_KEY]

    def get_total_length(self):
        return self.input_dict[CABLE_TYPE_TOTAL_LENGTH_KEY]

    def get_reel_length(self):
        return self.input_dict[CABLE_TYPE_REEL_LENGTH_KEY]

    def get_reel_qty(self):
        return self.input_dict[CABLE_TYPE_REEL_QTY_KEY]

    def get_lead_time(self):
        return self.input_dict[CABLE_TYPE_LEAD_TIME_KEY]

    def get_procurement_status(self):
        if len(self.input_dict[CABLE_TYPE_RECEIVED_KEY]) > 0:
            return "Received"
        elif len(self.input_dict[CABLE_TYPE_ORDERED_KEY]) > 0:
            return "Ordered"
        else:
            return "Unspecified"

    def get_procurement_info(self):
        proc_info = ""
        if len(str(self.input_dict[CABLE_TYPE_ORDERED_KEY])) > 0:
            proc_info = proc_info + "Ordered: " + str(self.input_dict[CABLE_TYPE_ORDERED_KEY]) + ". "
        if len(str(self.input_dict[CABLE_TYPE_RECEIVED_KEY])) > 0:
            proc_info = proc_info + "Received: " + str(self.input_dict[CABLE_TYPE_RECEIVED_KEY]) + "."

        return proc_info

    def get_project_id(self):
        return self.helper.get_project_id()

    def get_tech_system_id(self):
        return self.helper.get_tech_system_id()

    def get_owner_user_id(self):
        return self.helper.get_owner_user_id()

    def get_owner_group_id(self):
        return self.helper.get_owner_group_id()


class CableInventoryOutputObject(OutputObject):

    def __init__(self, helper, input_dict, tag_name, cable_type_name):
        super().__init__(helper, input_dict)
        self.tag_name = tag_name
        self.cable_type_name = cable_type_name

    @classmethod
    def get_output_columns(cls):
        column_list = [
            OutputColumnModel(method="empty_column", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_cable_type_name", label="Cable Catalog Item"),
            OutputColumnModel(method="get_tag", label="Tag"),
            OutputColumnModel(method="empty_column", label="QR ID"),
            OutputColumnModel(method="empty_column", label="Description"),
            OutputColumnModel(method="get_status", label="Status"),
            OutputColumnModel(method="empty_column", label="Location"),
            OutputColumnModel(method="empty_column", label="Location Details"),
            OutputColumnModel(method="empty_column", label="Length"),
            OutputColumnModel(method="get_project_id", label="Project"),
            OutputColumnModel(method="get_owner_user_id", label="Owner User"),
            OutputColumnModel(method="get_owner_group_id", label="Owner Group"),
        ]
        return column_list

    def get_cable_type_name(self):
        return "#" + self.cable_type_name

    def get_tag(self):
        return self.tag_name

    def get_status(self):
        return "Planned"

    def get_project_id(self):
        return self.helper.get_project_id()

    def get_owner_user_id(self):
        return self.helper.get_owner_user_id()

    def get_owner_group_id(self):
        return self.helper.get_owner_group_id()


class CableDesignOutputObject(OutputObject):

    ignore_port_columns = False

    def __init__(self, helper, input_dict, cable_catalog_info, ignore_port_columns, existing_item_id=""):
        super().__init__(helper, input_dict)
        self.cable_catalog_info = cable_catalog_info
        self.existing_item_id = existing_item_id

    @classmethod
    def get_output_columns(cls):

        if not cls.ignore_port_columns:
            endpoint1_port_method = "get_endpoint1_port"
            endpoint2_port_method = "get_endpoint2_port"
        else:
            endpoint1_port_method = "empty_column"
            endpoint2_port_method = "empty_column"

        column_list = [
            OutputColumnModel(method="get_existing_item_id", label="Existing Item ID"),
            OutputColumnModel(method="empty_column", label="Delete Existing Item"),
            OutputColumnModel(method="get_name", label="Name"),
            OutputColumnModel(method="get_cable_type", label="Type"),
            OutputColumnModel(method="empty_column", label="Assigned Inventory Tag"),
            OutputColumnModel(method="empty_column", label="Is Installed"),
            OutputColumnModel(method="get_alt_name", label="Alt Name"),
            OutputColumnModel(method="get_ext_name", label="Ext Cable Name"),
            OutputColumnModel(method="get_import_id", label="Import Cable ID"),
            OutputColumnModel(method="empty_column", label="Alternate Cable ID"),
            OutputColumnModel(method="empty_column", label="Description"),
            OutputColumnModel(method="get_laying", label="Laying"),
            OutputColumnModel(method="get_voltage", label="Voltage"),
            OutputColumnModel(method="empty_column", label="Routed Length (ft)"),
            OutputColumnModel(method="empty_column", label="Route"),
            OutputColumnModel(method="empty_column", label="Total Required Cable Length (ft)"),
            OutputColumnModel(method="empty_column", label="Notes"),
            OutputColumnModel(method="get_endpoint1_id", label="Endpoint1"),
            OutputColumnModel(method=endpoint1_port_method, label="Endpoint1 Port"),
            OutputColumnModel(method="get_endpoint1_connector", label="Endpoint1 Connector"),
            OutputColumnModel(method="get_endpoint1_description", label="Endpoint1 Desc"),
            OutputColumnModel(method="get_endpoint1_route", label="Endpoint1 Route"),
            OutputColumnModel(method="empty_column", label="Endpoint1 End Length (ft)"),
            OutputColumnModel(method="empty_column", label="Endpoint1 Termination"),
            OutputColumnModel(method="empty_column", label="Endpoint1 Pinlist"),
            OutputColumnModel(method="empty_column", label="Endpoint1 Notes"),
            OutputColumnModel(method="empty_column", label="Endpoint1 Drawing"),
            OutputColumnModel(method="get_endpoint2_id", label="Endpoint2"),
            OutputColumnModel(method=endpoint2_port_method, label="Endpoint2 Port"),
            OutputColumnModel(method="get_endpoint2_connector", label="Endpoint2 Connector"),
            OutputColumnModel(method="get_endpoint2_description", label="Endpoint2 Desc"),
            OutputColumnModel(method="get_endpoint2_route", label="Endpoint2 Route"),
            OutputColumnModel(method="empty_column", label="Endpoint2 End Length (ft)"),
            OutputColumnModel(method="empty_column", label="Endpoint2 Termination"),
            OutputColumnModel(method="empty_column", label="Endpoint2 Pinlist"),
            OutputColumnModel(method="empty_column", label="Endpoint2 Notes"),
            OutputColumnModel(method="empty_column", label="Endpoint2 Drawing"),
            OutputColumnModel(method="get_project_id", label="Project"),
            OutputColumnModel(method="get_tech_system_id", label="Technical System"),
            OutputColumnModel(method="get_owner_user_id", label="Owner User"),
            OutputColumnModel(method="get_owner_group_id", label="Owner Group"),
        ]
        return column_list

    def get_existing_item_id(self):
        return self.existing_item_id

    @classmethod
    def get_name_cls(cls, row_dict, cable_id=None, import_id=None):

        if import_id is not None:
            import_id = str(import_id)

        if cable_id is None and import_id is None:
            if row_dict is not None:
                cable_id = row_dict[CABLE_DESIGN_CABLE_ID_KEY]
                import_id = str(row_dict[CABLE_DESIGN_IMPORT_ID_KEY])
            else:
                return ""

        if cable_id is not None and cable_id != "":
            return cable_id
        else:
            if import_id is None:
                import_id = ""
            return "CA " + import_id

    def get_name(self):
        return self.get_name_cls(self.input_dict)

    def get_alt_name(self):
        return "<" + self.input_dict[CABLE_DESIGN_SRC_ETPM_KEY] + "><" + \
               self.input_dict[CABLE_DESIGN_DEST_ETPM_KEY] + ">:" + \
               self.get_name()

    def get_ext_name(self):
        return self.input_dict[CABLE_DESIGN_NAME_KEY]

    def get_import_id(self):
        return str(self.input_dict[CABLE_DESIGN_IMPORT_ID_KEY])

    def get_laying(self):
        return self.input_dict[CABLE_DESIGN_LAYING_KEY]

    def get_voltage(self):
        return self.input_dict[CABLE_DESIGN_VOLTAGE_KEY]

    def get_cable_type(self):
        cable_type_name = self.input_dict[CABLE_DESIGN_TYPE_KEY]
        if cable_type_name is not None and len(cable_type_name) > 0:
            return "#" + cable_type_name
        else:
            return None

    def get_endpoint_id(self, input_column_key, container_key):
        endpoint_name = self.input_dict[input_column_key]
        container_name = self.input_dict[container_key]
        return self.helper.info_manager.rack_manager.get_endpoint_id_for_rack(container_name, endpoint_name)

    def get_endpoint1_id(self):
        return self.get_endpoint_id(CABLE_DESIGN_END1_DEVICE_NAME_KEY, CABLE_DESIGN_SRC_ETPM_KEY)

    def get_endpoint1_port(self):
        return self.input_dict[CABLE_DESIGN_END1_PORT_NAME_KEY]

    def get_endpoint1_connector(self):
        if CABLE_TYPE_E1_1_KEY in self.cable_catalog_info.connector_names:
            return CABLE_TYPE_E1_1_KEY
        else:
            return None

    def get_endpoint1_description(self):
        return str(self.input_dict[CABLE_DESIGN_SRC_LOCATION_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_SRC_ANS_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_SRC_ETPM_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_SRC_ADDRESS_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_SRC_DESCRIPTION_KEY])

    def get_endpoint1_route(self):
        return str(self.input_dict[CABLE_DESIGN_VIA_ROUTE_KEY])

    def get_endpoint2_id(self):
        return self.get_endpoint_id(CABLE_DESIGN_END2_DEVICE_NAME_KEY, CABLE_DESIGN_DEST_ETPM_KEY)

    def get_endpoint2_port(self):
        return self.input_dict[CABLE_DESIGN_END2_PORT_NAME_KEY]

    def get_endpoint2_connector(self):
        if CABLE_TYPE_E2_1_KEY in self.cable_catalog_info.connector_names:
            return CABLE_TYPE_E2_1_KEY
        else:
            return None

    def get_endpoint2_description(self):
        return str(self.input_dict[CABLE_DESIGN_DEST_LOCATION_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_DEST_ANS_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_DEST_ETPM_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_DEST_ADDRESS_KEY]) + ":" + \
               str(self.input_dict[CABLE_DESIGN_DEST_DESCRIPTION_KEY])

    def get_endpoint2_route(self):
        return str(self.input_dict[CABLE_DESIGN_WAYPOINT_ROUTE_KEY])

    def get_project_id(self):
        return self.helper.get_project_id()

    def get_tech_system_id(self):
        return self.helper.get_tech_system_id()

    def get_owner_user_id(self):
        return self.helper.get_owner_user_id()

    def get_owner_group_id(self):
        return self.helper.get_owner_group_id()


class InputSheetHelper(ABC):

    def __init__(self):
        self.input_columns = {}
        self.input_column_key_index_dict = {}
        self.output_columns = {}
        self.input_handlers = []
        self.output_objects = []
        self.validation_map = {}
        self.info_manager = None
        self.api = None
        self.validate_only = False
        self.max_row = 0
        self.max_column = 0
        self.header_row = None
        self.first_data_row = None
        self.last_data_row = None
        self.input_sheet = None
        self.input_valid_message = None
        self.num_input_rows = None
        self.config_preimport = None
        self.config_group = None
        self.num_empty_rows = 0
        self.num_rows_missing_required_column = 0
        self.num_handler_validation_errors = 0
        self.num_rows_cell_parsed_date = 0
        self.num_comment_rows = 0

    # returns number of rows at which progress message should be displayed
    @classmethod
    def progress_increment(cls):
        return 5

    def post_create(self, config_preimport, config_workbook, info_manager, api):
        self.set_config_preimport(config_preimport, self.sheet_name())
        self.set_config_workbook(config_workbook, self.sheet_name())
        self.info_manager = info_manager
        self.api = api

    # Returns registered tag for subclass.
    @staticmethod
    @abstractmethod  # must be innermost decorator
    def sheet_name():
        pass

    # Returns 1-based sheet index for helper's sheet.
    @staticmethod
    @abstractmethod # must be innermost decorator
    def sheet_number():
        pass

    # Returns expected number of columns in input spreadsheet.
    def num_input_cols(self):
        return len(self.input_columns)

    def num_output_cols(self):
        return len(self.output_column_list())

    def pre_initialize(self, api, input_book):
        self.pre_initialize_custom(api, input_book)

    def pre_initialize_custom(self, api, input_book):
        pass

    # Initializes the helper.  Subclass overrides initialize_custom() to customize.
    def initialize(self, api, sheet, first_row, last_row):
        self.initialize_columns()
        self.initialize_handlers(self.info_manager, api, sheet, first_row, last_row)
        self.initialize_custom(api, sheet, first_row, last_row)

    def initialize_columns(self):
        self.initialize_input_columns()
        self.initialize_output_columns()

    # Builds dictionary whose keys are column index and value is column model object.
    def initialize_input_columns(self):
        column_index = 0
        for col in self.generate_input_column_list():
            self.input_columns[column_index] = col
            self.input_column_key_index_dict[col.key] = column_index
            column_index = column_index + 1

    def get_input_column_index_for_key(self, key):
        return self.input_column_key_index_dict[key]

    # Builds dictionary whose keys are column index and value is column model object.
    def initialize_output_columns(self):
        column_index = 0
        for col in self.generate_output_column_list():
            self.output_columns[column_index] = col
            column_index = column_index + 1

    def initialize_handlers(self, info_manager, api, sheet, first_row, last_row):
        for handler in self.generate_handler_list():
            self.input_handlers.append(handler)
            handler.info_manager = info_manager
            handler.initialize(api, sheet, first_row, last_row)

    def initialize_custom(self, api, sheet, first_row, last_row):
        pass

    # subclass overrides to create list of input columns
    @abstractmethod
    def generate_input_column_list(self):
        pass

    # subclass overrides to create list of output columns
    @abstractmethod
    def generate_output_column_list(self):
        pass

    # subclass overrides to create list of input handlers
    def generate_handler_list(self):
        pass

    # Returns list of input column models.
    def input_column_list(self):
        return list(self.input_columns.values())

    # Returns list of output column models.  Not all columns need be mapped, only the ones we wish to
    # write values to.
    def output_column_list(self):
        return list(self.output_columns.values())

    # Returns list of input handlers.
    def input_handler_list(self):
        return self.input_handlers

    # Returns an output object for the specified input object, or None if the input object is duplicate.
    @abstractmethod
    def handle_valid_row(self, input_dict):
        pass

    def set_config_preimport(self, config, section):
        self.config_preimport = config

    def set_config_workbook(self, config, section):

        self.config_group = config
        self.validate_only = get_config_resource_boolean(config, section, 'validateOnly', False)
        self.header_row = get_config_resource_int(config, section, 'headerRow', False)
        self.first_data_row = get_config_resource_int(config, section, 'firstDataRow', False)
        self.last_data_row = get_config_resource_int(config, section, 'lastDataRow', False)

    def get_header_row(self):
        return self.header_row

    def get_first_data_row(self):
        return self.first_data_row

    def get_last_data_row(self):
        return self.last_data_row

    def set_api(self, api):
        self.api = api

    def set_num_input_rows(self, num_input_rows):
        self.num_input_rows = num_input_rows

    def set_input_sheet(self, sheet):
        self.input_sheet = sheet

    def process_input_book(self, input_book):

        item_type = self.sheet_name().upper()

        # pre-initialize helper (allows it to access arbitraru data from workbook
        # without sheet and first/last row number
        self.pre_initialize(self.api, input_book)

        # process sheetNumber option
        # option_sheet_number = get_config_resource(self.config_preimport, self.sheet_name(), 'sheetNumber', True)

        sheet_num = self.sheet_number()
        sheet_index = sheet_num - 1
        input_sheet = input_book.worksheets[int(sheet_index)]
        self.max_row = input_sheet.max_row
        self.max_column = input_sheet.max_column
        logging.info("input spreadsheet dimensions: %d x %d" % (input_sheet.max_row, input_sheet.max_column))

        header_row_num = self.get_header_row()
        first_data_row_num = self.get_first_data_row()
        last_data_row_num = self.get_last_data_row()
        header_index = header_row_num
        first_data_index = first_data_row_num
        last_data_index = last_data_row_num

        # validate input spreadsheet dimensions
        if header_index == 0 or first_data_index == 0 or last_data_index == 0:
            fatal_error("data row index values cannot be zero header: %d first data: %d last data: %d"
                        % (header_index, first_data_index, last_data_index))
        if input_sheet.max_row < last_data_row_num:
            fatal_error("fewer rows in input sheet than last data row: %d, exiting" % last_data_row_num)
        if input_sheet.max_column < self.num_input_cols():
            fatal_error("input sheet actual columns: %d less than expected columns: %d, exiting" % (
                input_sheet.max_column, self.num_input_cols()))

        # initialize helper
        print()
        print("%s: INITIALIZING AND FETCHING CDB DATA ====================" % item_type)
        print()
        self.set_input_sheet(input_sheet)
        self.initialize(self.api, input_sheet, first_data_index, last_data_index)

        #
        # process rows from input spreadsheet
        #

        print()
        print("%s: PROCESSING SPREADSHEET ROWS ====================" % item_type)
        print("first data row: %d, last data row: %d" % (first_data_index, last_data_index))
        print()

        input_valid = True
        num_input_rows = 0
        rows = input_sheet.rows
        row_ind = 0

        # validate header and first data indexes
        if header_index < row_ind:
            fatal_error("invalid header_index: %d" % header_index)
        if first_data_index <= header_index:
            fatal_error(
                "invalid first data row: %d less than expected header row: %d" % (first_data_index, header_index))

        for row in rows:

            row_ind = row_ind + 1

            # skip to header row
            if row_ind < header_index:
                continue

            # validate header row
            elif row_ind == header_index:
                ignored_columns = []
                num_header_cols = 0
                header_cell_ind = 0
                for cell in row:
                    header_cell_value = cell.value
                    if header_cell_ind in self.input_columns:
                        header_input_column = self.input_columns[header_cell_ind]
                        if header_input_column is None:
                            fatal_error(
                                "unexpected actual header column: %s index: %d" % (header_cell_value, header_cell_ind))
                        expected_header_label = header_input_column.label
                        if expected_header_label is not None:
                            # ignore mismatch when expected value not specified (for cases like CableSpecs where the header value changes for each tech system
                            if header_cell_value != expected_header_label:
                                fatal_error("actual header column: %s mismatch with expected: %s" % (
                                    header_cell_value, expected_header_label))
                    else:
                        if header_cell_value is not None and header_cell_value != '':
                            ignored_columns.append(header_cell_value)
                    num_header_cols = num_header_cols + 1
                    header_cell_ind = header_cell_ind + 1
                    continue
                if len(ignored_columns) > 0:
                    print("ignored extra input columns: %s" % ignored_columns)

            # skip to first data row:
            elif row_ind < first_data_index:
                continue

            # skip trailing rows
            elif row_ind > last_data_index:
                continue

            # process data rows
            else:

                current_row_num = row_ind
                num_input_rows = num_input_rows + 1

                logging.debug("processing row %d from input spreadsheet" % current_row_num)

                input_dict = {}
                row_is_valid = True
                row_valid_messages = []
                row_error_parsed_date = False
                row_is_comment = False

                for col_ind in range(self.num_input_cols()):

                    # read cell value from spreadsheet
                    val = row[col_ind].value
                    if val is None:
                        val = ""
                    if isinstance(val, datetime):
                        # we don't want to parse anything as a date from the input sheet, this happens when a value
                        # like rack name "03-25" gets turned into a date like "03/25/2022 00:00:00" by excel
                        row_is_valid = False
                        row_valid_messages.append("parsed as date from excel input sheet: %s" % str(val))
                        val = str(val)
                        row_error_parsed_date = True

                    # check for comment row
                    if col_ind == 0 and isinstance(val, str) and val.startswith("//"):
                        self.num_comment_rows = self.num_comment_rows + 1
                        row_is_comment = True
                        break

                    # parse expected columns
                    if col_ind in self.input_columns:
                        logging.debug("col: %d value: %s" % (col_ind, str(val)))
                        self.handle_input_cell_value(input_dict=input_dict, index=col_ind, value=val,
                                                     row_num=current_row_num)

                # skip further processing of comment rows
                if row_is_comment:
                    continue

                # count date parsing errors for display on error sheet
                if row_error_parsed_date:
                    self.num_rows_cell_parsed_date = self.num_rows_cell_parsed_date + 1

                # ignore row if blank
                if self.input_row_is_empty(input_dict=input_dict, row_num=current_row_num):
                    self.num_empty_rows = self.num_empty_rows + 1
                    continue

                # validate row
                (is_valid, valid_messages) = self.input_row_is_valid(input_dict=input_dict, row_num=row_ind)
                if not is_valid:
                    row_is_valid = False
                    row_valid_messages.extend(valid_messages)

                # invoke handlers
                (handler_is_valid, handler_messages) = self.invoke_row_handlers(input_dict=input_dict,
                                                                                row_num=row_ind)
                if not handler_is_valid:
                    self.num_handler_validation_errors = self.num_handler_validation_errors + 1
                    row_is_valid = False
                    row_valid_messages.extend(handler_messages)

                if row_is_valid:
                    self.handle_valid_row(input_dict=input_dict)
                else:
                    input_valid = False
                    msg = "ERRORS found for row %d" % current_row_num
                    logging.error(msg)
                    self.validation_map[current_row_num] = row_valid_messages

                # print progress message
                if num_input_rows % self.progress_increment() == 0:
                    print("processed %d spreadsheet rows" % num_input_rows, flush=True)

        # print final progress message
        print("processed %d spreadsheet rows" % num_input_rows)
        self.set_num_input_rows(num_input_rows)

        #
        # display validation summary
        #

        print()
        print("%s: VALIDATION SUMMARY ====================" % item_type)
        print()

        (sheet_valid, sheet_valid_string) = self.input_is_valid()
        if not sheet_valid:
            input_valid = False
            msg = "ERROR validating input spreadsheet: %s" % sheet_valid_string
            logging.error(msg)
            print(msg)

        elif len(self.validation_map) > 0:
            print("%d validation ERROR(S) found" % len(self.validation_map))
            print("See output workbook for additional details")

        else:
            print("no validation errors found")

        if input_valid and not self.validate_only:
            summary_msg = "PROCESSING SUCCESSFUL: processed %d input rows including %d empty rows and wrote %d output rows, see output workbook for CDB import sheets\n" % (
                num_input_rows, self.num_empty_rows, len(self.output_objects))
            summary_msg = summary_msg + self.get_processing_summary()

        elif not input_valid:
            summary_msg = "PROCESSING ERROR: processed %d input rows including %d empty rows but no CDB import sheets generated, see output workbook and log for details" % (
                num_input_rows, self.num_empty_rows)

        else:
            summary_msg = "VALIDATION ONLY: processed %d input rows including %d empty rows but no CDB import sheets generated, see output workbook for details" % (
                num_input_rows, self.num_empty_rows)

        #
        # print summary
        #

        print()
        print("%s: PROCESSING SUMMARY ====================" % item_type)
        print()

        print(summary_msg)
        logging.info(summary_msg)

        return input_valid

    # Handles cell value from input spreadsheet at specified column index for supplied input object.
    def handle_input_cell_value(self, input_dict, index, value, row_num):
        key = self.input_columns[index].key
        input_dict[key] = value

    def input_row_is_empty(self, input_dict, row_num):

        non_empty_cols = {k: v for k, v in input_dict.items() if v is not None and v != ""}
        non_empty_count = len(non_empty_cols)
        blank_row_columns_allowed_dict = self.blank_row_columns_allowed_dict()

        # row contains no values
        if non_empty_count == 0:
            return True

        if self.input_row_is_empty_custom(input_dict, row_num):
            return True

        # check if values in more columns than allowed columns
        if non_empty_count > len(blank_row_columns_allowed_dict):
            return False

        # check if value in any column not explicitly allowed
        for column_key in non_empty_cols.keys():

            if column_key not in blank_row_columns_allowed_dict:
                return False

            allowed_value = blank_row_columns_allowed_dict[column_key]
            if allowed_value is not None:
                if non_empty_cols[column_key] != allowed_value:
                    return False

        return True

    # Returns dictionary of column keys that can contain values and the row still be considered blank.
    # Dictionary values should be None, or a string that is the value allowed in that column if we want that constraint.
    # Subclasses override to customize
    @classmethod
    def blank_row_columns_allowed_dict(cls):
        return {}

    # Returns True if the row represented by input_dict should be treated as a blank row.  Default is False.  Subclass
    # can override to allow certain non-empty values to be treated as empty.
    def input_row_is_empty_custom(self, input_dict, row_num):
        return False

    # Performs validation on row from input spreadsheet and returns True if the row is determined to be valid.
    # Can return False where input is valid, but it might be better to call sys.exit() with a useful message.
    def input_row_is_valid(self, input_dict, row_num):

        is_valid = True
        valid_messages = []

        missing_required_column = False
        for column in self.input_column_list():

            required = column.required
            if required:
                value = input_dict[column.key]
                if value is None or len(str(value)) == 0:
                    is_valid = False
                    valid_messages.append("required value missing for key: %s row index: %d" % (column.key, row_num))
                    missing_required_column = True

        if missing_required_column:
            self.num_rows_missing_required_column = self.num_rows_missing_required_column + 1

        (custom_is_valid, custom_valid_string) = self.input_row_is_valid_custom(input_dict)
        if not custom_is_valid:
            is_valid = False
            valid_messages.append(custom_valid_string)

        return is_valid, valid_messages

    # Performs custom validation on input row.  Returns True if row is valid.  Default is to return True. Subclass
    # can override to customize.
    def input_row_is_valid_custom(self, input_dict):
        return True, ""

    # Provides hook for subclasses to override to validate the input before generating the output spreadsheet.
    def input_is_valid(self):
        return True, ""

    def invoke_row_handlers(self, input_dict, row_num):

        is_valid = True
        valid_messages = []

        for handler in self.input_handler_list():
            (handler_is_valid, handler_valid_string) = handler.handle_input(input_dict)
            if not handler_is_valid:
                is_valid = False
                valid_messages.append(handler_valid_string)

        return is_valid, valid_messages

    # Returns column label for specified column index.
    def get_output_column_label(self, col_index):
        return self.output_columns[col_index].label

    def get_cell_value(self, obj, method):
        # use reflection to invoke column getter method on supplied object
        val = getattr(obj, method)()
        return val

    def write_sheet(self, output_book, sheet_name, columns, output_objects):

        output_sheet = output_book.add_worksheet(sheet_name)

        # write output spreadsheet header row
        row_ind = 0
        col_ind = 0
        for column in columns:
            label = column.label
            output_sheet.write(row_ind, col_ind, label)
            col_ind = col_ind + 1

        # process rows
        num_output_rows = 0
        for output_obj in output_objects:

            row_ind = row_ind + 1
            num_output_rows = num_output_rows + 1
            current_row_num = row_ind + 1

            col_ind = 0
            for column in columns:
                val = self.get_cell_value(output_obj, column.method)
                output_sheet.write(row_ind, col_ind, val)
                col_ind = col_ind + 1

        return num_output_rows

    def get_summary_messages(self):
        return ["Total rows processed in input sheet: %d" % self.num_input_rows,
                "Blank rows in input sheet: %d" % self.num_empty_rows,
                "Comment rows in input sheet: %d" % self.num_comment_rows]

    # Subclasses override to return list of custom summary messages for summary sheet
    def get_summary_messages_custom(self):
        return []

    def get_summary_sheet_columns(self):
        return [], []

    def write_summary_sheet(self, output_book):

        summary_messages = self.get_summary_messages()
        summary_messages.extend(self.get_summary_messages_custom())
        summary_column_names, summary_column_values = self.get_summary_sheet_columns()

        if len(summary_messages) == 0 and len(summary_column_names) == 0:
            return

        output_sheet = output_book.add_worksheet(self.sheet_name() + " Sheet Summary")
        text_wrap_format = output_book.add_format({'text_wrap': True})

        column_index = 0
        if len(summary_messages) > 0:

            output_sheet.write(0, 0, "summary messages")
            output_sheet.set_column(0, 0, DEFAULT_COLUMN_WIDTH)

            row_index = 1
            for summary_message in summary_messages:
                row_height = get_row_height_wrapped_message(summary_message)
                output_sheet.write(row_index, 0, summary_message + "\n", text_wrap_format)
                output_sheet.set_row(row_index, row_height)
                row_index = row_index + 1

            column_index = column_index + 1

        if len(summary_column_names) > 0:

            for (column_name, column_values) in zip(summary_column_names, summary_column_values):
                maximum_column_width = 0
                row_index = 0
                # write column header
                output_sheet.write(row_index, column_index, column_name)
                if len(column_name) > maximum_column_width:
                    maximum_column_width = len(column_name)
                for value in column_values:
                    row_index = row_index + 1
                    output_sheet.write(row_index, column_index, value)
                    if len(value) > maximum_column_width:
                        maximum_column_width = len(value)

                # set column width
                output_sheet.set_column(column_index, column_index, maximum_column_width)

                column_index = column_index + 1

    # Writes content to output workbook.  Subclasses override to customize with multiple tabs
    @abstractmethod
    def write_helper_sheets(self, output_book):
        pass

    def write_workbook_sheets(self, output_book):
        self.write_summary_sheet(output_book)
        self.write_helper_sheets(output_book)

    def get_validate_only_messages(self):
        return []

    def get_validate_only_sheet_columns(self):
        return [], []

    def write_validate_only_sheet(self, output_book):

        validate_only_messages = self.get_validate_only_messages()
        validate_only_column_names, validate_only_column_values = self.get_validate_only_sheet_columns()

        validate_only_messages.insert(0, "No validation errors in input sheet")
        validate_only_messages.insert(0, "Data rows in input sheet: %d" % self.num_input_rows)

        if len(validate_only_messages) == 0 and len(validate_only_column_names) == 0:
            return

        output_sheet = output_book.add_worksheet(self.sheet_name() + " Sheet Validation")
        text_wrap_format = output_book.add_format({'text_wrap': True})

        column_index = 0
        if len(validate_only_messages) > 0:

            output_sheet.write(0, 0, "validate only messages")
            output_sheet.set_column(0, 0, DEFAULT_COLUMN_WIDTH)

            row_index = 1
            for summary_message in validate_only_messages:
                row_height = get_row_height_wrapped_message(summary_message)
                output_sheet.write(row_index, 0, summary_message + "\n", text_wrap_format)
                output_sheet.set_row(row_index, row_height)
                row_index = row_index + 1

            column_index = column_index + 1

        if len(validate_only_column_names) > 0:

            for (column_name, column_values) in zip(validate_only_column_names, validate_only_column_values):
                maximum_column_width = 0
                row_index = 0
                # write column header
                output_sheet.write(row_index, column_index, column_name)
                if len(column_name) > maximum_column_width:
                    maximum_column_width = len(column_name)
                for value in column_values:
                    row_index = row_index + 1
                    output_sheet.write(row_index, column_index, value)
                    if len(value) > maximum_column_width:
                        maximum_column_width = len(value)

                # set column width
                output_sheet.set_column(column_index, column_index, maximum_column_width)

                column_index = column_index + 1

    def write_validate_only_sheets(self, output_book):
        self.write_validate_only_sheet(output_book)

    # Returns value for output spreadsheet cell and supplied object at specified index.
    def get_output_cell_value(self, obj, index):
        # use reflection to invoke column getter method on supplied object
        val = getattr(obj, self.output_columns[index].method)()
        logging.debug("index: %d method: %s value: %s" % (index, self.output_columns[index].method, val))
        return val

    def get_error_messages(self):
        return ["Rows missing required values: %d" % self.num_rows_missing_required_column,
                "Rows with handler validation errors: %d" % self.num_handler_validation_errors,
                "Rows with cells parsed as Excel date: %d" % self.num_rows_cell_parsed_date]

    def get_error_messages_custom(self):
        return []

    def get_error_sheet_columns(self):
        return [], []

    def write_error_sheet(self, output_book):

        error_messages = self.get_error_messages()
        error_messages.extend(self.get_error_messages_custom())
        error_column_names, error_column_values = self.get_error_sheet_columns()

        if self.input_valid_message is None and len(error_messages) == 0 and len(error_column_names) == 0:
            return

        output_sheet = output_book.add_worksheet(self.sheet_name() + " " + "Sheet Errors")
        text_wrap_format = output_book.add_format({'text_wrap': True})

        column_index = 0
        if self.input_valid_message is not None or len(error_messages) > 0:

            output_sheet.write(0, 0, "error messages")
            output_sheet.set_column(0, 0, DEFAULT_COLUMN_WIDTH)

            row_index = 1
            if self.input_valid_message is not None:
                row_height = get_row_height_wrapped_message(self.input_valid_message)
                output_sheet.write(row_index, 0, self.input_valid_message + "\n", text_wrap_format)
                output_sheet.set_row(row_index, row_height)
                row_index = row_index + 1

            for error_message in error_messages:
                row_height = get_row_height_wrapped_message(error_message)
                output_sheet.write(row_index, 0, error_message + "\n", text_wrap_format)
                output_sheet.set_row(row_index, row_height)
                row_index = row_index + 1

            column_index = column_index + 1

        if len(error_column_names) > 0:

            for (column_name, column_values) in zip(error_column_names, error_column_values):
                maximum_column_width = 0
                row_index = 0
                # write column header
                output_sheet.write(row_index, column_index, column_name)
                if len(column_name) > maximum_column_width:
                    maximum_column_width = len(column_name)
                for value in column_values:
                    row_index = row_index + 1
                    output_sheet.write(row_index, column_index, value)
                    if len(value) > maximum_column_width:
                        maximum_column_width = len(value)

                # set column width
                output_sheet.set_column(column_index, column_index, maximum_column_width)

                column_index = column_index + 1

    # Writes error content to output workbook, subclasses override to customize with multiple error sheets.
    def write_error_sheets(self, output_book):
        self.write_error_sheet(output_book)

    def write_validation_sheet(self, output_book):

        if len(self.validation_map) == 0:
            return

        # create output sheet
        output_sheet = output_book.add_worksheet(self.sheet_name() + " " + "Sheet Row Errors")

        # write header row
        output_sheet.write(0, 0, "input row number")
        output_sheet.write(0, 1, "validation messages")

        row_ind = 0
        maximum_column_width = 0
        for key in self.validation_map:
            row_ind = row_ind + 1
            output_sheet.write(row_ind, 0, key)
            messages = ""
            message_count = len(self.validation_map[key])
            for message in self.validation_map[key]:
                messages = messages + message + "\n"
                if len(message) > maximum_column_width:
                    maximum_column_width = len(message)
            output_sheet.write(row_ind, 1, messages)
            # set row height
            output_sheet.set_row(row_ind, (message_count + 1) * DEFAULT_FONT_HEIGHT)

        # set column width
        output_sheet.set_column(1, 1, maximum_column_width)

    # Returns processing summary message.
    def get_processing_summary(self):
        return ""


class CableSpecsSheetHelper(InputSheetHelper):

    def __init__(self):
        super().__init__()
        self.project_id = None
        self.tech_system_id = None
        self.owner_user_id = None
        self.owner_group_id = None
        self.workbook_tech_group = None

    def get_project_id(self):
        return self.project_id

    def get_tech_system_id(self):
        return self.tech_system_id

    def get_owner_user_id(self):
        return self.owner_user_id

    def get_owner_group_id(self):
        return self.owner_group_id

    def set_config_preimport(self, config, section):
        super().set_config_preimport(config, section)
        self.project_id = get_config_resource(config, section, 'cdbProjectId', True)
        self.owner_user_id = get_config_resource(config, section, 'cdbOwnerUserId', True)
        self.owner_group_id = get_config_resource(config, section, 'cdbOwnerGroupId', True)

    def set_config_workbook(self, config, section):
        super().set_config_workbook(config, section)
        self.tech_system_id = get_config_resource(config, section, 'cdbTechSystemId', True)
        self.workbook_tech_group = get_config_resource(config, section, 'kabelWorkbookTechGroup', True)

    @staticmethod
    def sheet_name():
        return "CableSpecs"

    @staticmethod
    def sheet_number():
        return 3

    @staticmethod
    def item_name():
        return "Cable Catalog"

    def generate_input_column_list(self):
        column_list = [
            InputColumnModel(key=CABLE_TYPE_NAME_KEY, required=True),
            InputColumnModel(key=CABLE_TYPE_DESCRIPTION_KEY, label=LABEL_CABLESPECS_DESCRIPTION),
            InputColumnModel(key=CABLE_TYPE_MANUFACTURER_KEY, label=LABEL_CABLESPECS_MANUFACTURER),
            InputColumnModel(key=CABLE_TYPE_PART_NUMBER_KEY, label=LABEL_CABLESPECS_PART_NUM),
            InputColumnModel(key=CABLE_TYPE_ALT_PART_NUMBER_KEY, label=LABEL_CABLESPECS_ALT_PART_NUM),
            InputColumnModel(key=CABLE_TYPE_DIAMETER_KEY, label=LABEL_CABLESPECS_DIAMETER),
            InputColumnModel(key=CABLE_TYPE_WEIGHT_KEY, label=LABEL_CABLESPECS_WEIGHT),
            InputColumnModel(key=CABLE_TYPE_CONDUCTORS_KEY, label=LABEL_CABLESPECS_CONDUCTORS),
            InputColumnModel(key=CABLE_TYPE_INSULATION_KEY, label=LABEL_CABLESPECS_INSULATION),
            InputColumnModel(key=CABLE_TYPE_JACKET_COLOR_KEY, label=LABEL_CABLESPECS_JACKET),
            InputColumnModel(key=CABLE_TYPE_VOLTAGE_RATING_KEY, label=LABEL_CABLESPECS_VOLTAGE),
            InputColumnModel(key=CABLE_TYPE_FIRE_LOAD_KEY, label=LABEL_CABLESPECS_FIRE_LOAD),
            InputColumnModel(key=CABLE_TYPE_HEAT_LIMIT_KEY, label=LABEL_CABLESPECS_HEAT_LIMIT),
            InputColumnModel(key=CABLE_TYPE_BEND_RADIUS_KEY, label=LABEL_CABLESPECS_BEND_RADIUS),
            InputColumnModel(key=CABLE_TYPE_RAD_TOLERANCE_KEY, label=LABEL_CABLESPECS_RAD_TOLERANCE),
            InputColumnModel(key=CABLE_TYPE_LINK_URL_KEY, label=LABEL_CABLESPECS_LINK),
            InputColumnModel(key=CABLE_TYPE_IMAGE_URL_KEY, label=LABEL_CABLESPECS_IMAGE),
            InputColumnModel(key=CABLE_TYPE_TOTAL_LENGTH_KEY, label=LABEL_CABLESPECS_TOTAL_LENGTH),
            InputColumnModel(key=CABLE_TYPE_REEL_LENGTH_KEY, label=LABEL_CABLESPECS_REEL_LENGTH),
            InputColumnModel(key=CABLE_TYPE_REEL_QTY_KEY, label=LABEL_CABLESPECS_REEL_QUANTITY),
            InputColumnModel(key=CABLE_TYPE_LEAD_TIME_KEY, label=LABEL_CABLESPECS_LEAD_TIME),
            InputColumnModel(key=CABLE_TYPE_ORDERED_KEY, label=LABEL_CABLESPECS_ORDERED),
            InputColumnModel(key=CABLE_TYPE_RECEIVED_KEY, label=LABEL_CABLESPECS_RECEIVED),
            InputColumnModel(key=CABLE_TYPE_CHECKLIST_KEY, label=LABEL_CABLESPECS_CHECKLIST),
            InputColumnModel(key=CABLE_TYPE_COLUMN_Y_KEY),
            InputColumnModel(key=CABLE_TYPE_COLUMN_Z_KEY),
            InputColumnModel(key=CABLE_TYPE_E1_1_KEY, label=LABEL_CABLESPECS_E1_1),
            InputColumnModel(key=CABLE_TYPE_E2_1_KEY, label=LABEL_CABLESPECS_E2_1),
            InputColumnModel(key=CABLE_TYPE_E1_2_KEY, label=LABEL_CABLESPECS_E1_2),
            InputColumnModel(key=CABLE_TYPE_E2_2_KEY, label=LABEL_CABLESPECS_E2_2),
            InputColumnModel(key=CABLE_TYPE_E1_3_KEY, label=LABEL_CABLESPECS_E1_3),
            InputColumnModel(key=CABLE_TYPE_E2_3_KEY, label=LABEL_CABLESPECS_E2_3),
        ]
        return column_list

    def generate_output_column_list(self):
        return CableTypeOutputObject.get_output_columns()

    def generate_handler_list(self):

        handler_list = [
            TechnicalSystemCableTypeValidationHandler(CABLE_TYPE_NAME_KEY, self.info_manager),
            UniqueNameHandler(CABLE_TYPE_NAME_KEY, self.info_manager.cable_type_names),
            CableTypeConnectorHandler(CABLE_TYPE_E1_1_KEY, 1),
            CableTypeConnectorHandler(CABLE_TYPE_E2_1_KEY, 2),
            CableTypeConnectorHandler(CABLE_TYPE_E1_2_KEY, 1),
            CableTypeConnectorHandler(CABLE_TYPE_E2_2_KEY, 2),
            CableTypeConnectorHandler(CABLE_TYPE_E1_3_KEY, 1),
            CableTypeConnectorHandler(CABLE_TYPE_E2_3_KEY, 2),
        ]

        if not self.validate_only:
            handler_list.append(CableTypeExistenceHandler(self.info_manager, CABLE_TYPE_NAME_KEY))
            mfr_index = self.get_input_column_index_for_key(CABLE_TYPE_MANUFACTURER_KEY)
            handler_list.append(SourceHandler(self.info_manager, CABLE_TYPE_MANUFACTURER_KEY, mfr_index, self.api))

        return handler_list

    def pre_initialize_custom(self, api, input_book):
        
        technical_system = self.workbook_tech_group
        self.info_manager.set_technical_system(technical_system)

        # find column for specified technical system in CableTypes tab
        cable_types_sheet = input_book.worksheets[1]
        max_row = cable_types_sheet.max_row
        max_column = cable_types_sheet.max_column
        row_ind = 1
        cable_type_column_ind = 0
        for col_ind in range(1, max_column+1):
            cell_value = cable_types_sheet.cell(row_ind, col_ind).value
            if technical_system == str(cell_value):
                cable_type_column_ind = col_ind
                break
        if cable_type_column_ind == 0:
            fatal_error("tech system: %s not found in CableTypes sheet row 1" % technical_system)

        # read cable types from appropriate column (used to know how many cable types to read from CableSpecs,
        # and also for validation that cable types in CableSpecs tab are valid
        cable_types = []
        for row_ind in range(11, max_row+1):
            cell_value = cable_types_sheet.cell(row_ind, cable_type_column_ind).value
            if cell_value is not None:
                cable_types.append(cell_value)
        if len(cable_types) == 0:
            fatal_error("no cable types found in CableTypes sheet column: %s" % technical_system)
        self.info_manager.set_cable_types_specified_for_technical_system(cable_types)

        # find header row for specified tech system in CableSpecs tab, set header, first, last data row numbers
        cable_specs_sheet = input_book.worksheets[2]
        max_row = cable_specs_sheet.max_row
        max_column = cable_specs_sheet.max_column
        col_ind = 1
        tech_system_row_ind = 0
        for row_ind in range(10, max_row+1):
            cell_value = cable_specs_sheet.cell(row_ind, col_ind).value
            if technical_system == str(cell_value):
                tech_system_row_ind = row_ind
                break
        if tech_system_row_ind == 0:
            fatal_error("header for tech system: %s not found in CableSpecs sheet column 1" % technical_system)
        self.header_row = tech_system_row_ind
        self.first_data_row = self.header_row + 1
        self.last_data_row = self.header_row + len(cable_types)

    def initialize_custom(self, api, sheet, first_row, last_row):

        # initialize connector type information from the connector columns
        connector_type_names = set()  # use set to eliminate duplicates
        for row_ind in range(first_row, last_row + 1):
            first_connector_type_index = self.get_input_column_index_for_key(CABLE_TYPE_E1_1_KEY)
            last_connector_type_index = self.get_input_column_index_for_key(CABLE_TYPE_E2_3_KEY)
            for col_ind in range(first_connector_type_index, last_connector_type_index+1):
                # get connector type name from all connector columns and add to set
                val = sheet.cell(row_ind, col_ind+1).value
                if val is not None and val != "":
                    connector_type_names.add(val)
        self.info_manager.initialize_connector_types(connector_type_names)

        # retrieve information for existing cable catalog items
        cable_type_names = set()  # use set to eliminate duplicates
        for row_ind in range(first_row, last_row+1):
            name_index = self.get_input_column_index_for_key(CABLE_TYPE_NAME_KEY)
            # cell() is 1-based so increment name index
            val = sheet.cell(row_ind, name_index+1).value
            if val is not None and val != "":
                cable_type_names.add(val)
        self.info_manager.load_cable_type_info(cable_type_names)

    def handle_valid_row(self, input_dict):

        name = input_dict[CABLE_TYPE_NAME_KEY]
        if name in self.info_manager.existing_cable_types:
            # add to list of output objects to generate tab for comparing existing cable types
            cable_type_id = self.info_manager.get_cable_type_id(name)
            compare_object = CableTypeOutputObject(helper=self, input_dict=input_dict, existing_item_id=cable_type_id)
            self.info_manager.output_objects_cable_type_compare.append(compare_object)
            return

        logging.debug("adding output object for: %s" % input_dict[CABLE_TYPE_NAME_KEY])
        output_object = CableTypeOutputObject(helper=self, input_dict=input_dict)
        self.output_objects.append(output_object)

    def input_is_valid(self):

        is_valid = True
        valid_string = ""

        # check that all cable types read from CableTypes tab are defined in the cable_specs tab
        for cable_type in self.info_manager.get_cable_types_specified_for_technical_system():
            if cable_type not in self.info_manager.get_cable_types_defined_for_technical_system():
                self.info_manager.undefined_cable_types.append(cable_type)
        if len(self.info_manager.undefined_cable_types) > 0:
            technical_system = self.info_manager.get_technical_system()
            is_valid = False
            valid_string = "not all cable types for technical system: %s specified in CableTypes tab are defined in CableSpecs tab: %s" % (technical_system, self.info_manager.undefined_cable_types)

        return is_valid, valid_string

    def get_summary_messages_custom(self):
        return ["Connector Types that already exist in CDB: %d" % len(self.info_manager.existing_connector_types),
                "Connector Types that need to be added to CDB: %d" % len(self.info_manager.new_connector_types),
                "Sources that already exist in CDB: %d" % len(self.info_manager.existing_sources),
                "Sources that need to be added to CDB: %d" % len(self.info_manager.new_sources),
                "Cable types that already exist in CDB: %d" % len(self.info_manager.existing_cable_types),
                "Cable types that need to be added to CDB: %d" % len(self.info_manager.new_cable_types)]

    def get_summary_sheet_columns(self):
        
        summary_column_names = []
        summary_column_values = []

        if len(self.info_manager.existing_connector_types) > 0:
            summary_column_names.append("existing connector types")
            summary_column_values.append(sorted(self.info_manager.existing_connector_types))

        if len(self.info_manager.new_connector_types) > 0:
            summary_column_names.append("new connector types")
            summary_column_values.append(sorted(self.info_manager.new_connector_types))

        if len(self.info_manager.existing_sources) > 0:
            summary_column_names.append("existing sources")
            summary_column_values.append(sorted(self.info_manager.existing_sources))

        if len(self.info_manager.new_sources) > 0:
            summary_column_names.append("new sources")
            summary_column_values.append(sorted(self.info_manager.new_sources))

        if len(self.info_manager.existing_cable_types) > 0:
            summary_column_names.append("existing cable types")
            summary_column_values.append(sorted(self.info_manager.existing_cable_types))

        if len(self.info_manager.new_cable_types) > 0:
            summary_column_names.append("new cable types")
            summary_column_values.append(sorted(self.info_manager.new_cable_types))

        return summary_column_names, summary_column_values

    def write_helper_sheets(self, output_book):

        self.write_sheet(output_book,
                         "Connector Type Import",
                         ConnectorTypeOutputObject.get_output_columns(),
                         self.info_manager.output_connector_types)

        self.write_sheet(output_book,
                         "Source Item Import",
                         SourceOutputObject.get_output_columns(),
                         self.info_manager.output_objects_source)

        self.write_sheet(output_book,
                         "Cable Catalog Item Compare",
                         CableTypeOutputObject.get_output_columns(),
                         self.info_manager.output_objects_cable_type_compare)

        self.write_sheet(output_book, "Cable Catalog Item Import", self.output_column_list(), self.output_objects)

        self.write_sheet(output_book,
                         "Cable Catalog Connector Import",
                         CableTypeConnectorOutputObject.get_output_columns(),
                         self.info_manager.output_cable_type_connectors)

    # Returns processing summary message.
    def get_processing_summary(self):
        if len(self.info_manager.existing_cable_types) > 0:
            return "DETAILS: number of cable types that already exist in CDB (not written to output file): %d" % len(self.info_manager.existing_cable_types)
        else:
            return ""

    def get_error_messages_custom(self):

        messages = []

        if len(self.info_manager.undefined_cable_types) > 0:
            messages.append("Cable types specified in CableTypes tab for technical system: %s not specified in CableSpecs tab: %d" % (self.info_manager.get_technical_system(), len(self.info_manager.undefined_cable_types)))

        return messages

    def get_error_sheet_columns(self):

        error_column_names = []
        error_column_values = []

        if len(self.info_manager.undefined_cable_types) > 0:
            error_column_names.append("cable types not defined in CableSpecs tab")
            error_column_values.append(sorted(self.info_manager.undefined_cable_types))

        return error_column_names, error_column_values


class CablesSheetHelper(InputSheetHelper):

    def __init__(self):
        super().__init__()
        self.md_root = None
        self.project_id = None
        self.tech_system_id = None
        self.owner_user_id = None
        self.owner_group_id = None
        self.ignore_port_columns = False

    def get_project_id(self):
        return self.project_id

    def get_tech_system_id(self):
        return self.tech_system_id

    def get_owner_user_id(self):
        return self.owner_user_id

    def get_owner_group_id(self):
        return self.owner_group_id

    # returns number of rows at which progress message should be displayed
    @classmethod
    def progress_increment(cls):
        return 50

    def set_config_preimport(self, config, section):
        super().set_config_preimport(config, section)
        self.project_id = get_config_resource(config, section, 'cdbProjectId', True)
        self.owner_user_id = get_config_resource(config, section, 'cdbOwnerUserId', True)
        self.owner_group_id = get_config_resource(config, section, 'cdbOwnerGroupId', True)
        self.md_root = get_config_resource(config, section, 'cdbMachineDesignRoot', True)

    def set_config_workbook(self, config, section):
        super().set_config_workbook(config, section)
        self.tech_system_id = get_config_resource(config, section, 'cdbTechSystemId', True)
        self.ignore_port_columns = get_config_resource_boolean(config, section, 'ignorePortColumns', False)

    @staticmethod
    def sheet_name():
        return "Cables"

    @staticmethod
    def sheet_number():
        return 1

    @staticmethod
    def item_name():
        return "Cable Design"

    def get_header_row(self):
        return 19

    def get_first_data_row(self):
        return 20

    def get_last_data_row(self):
        return self.max_row

    def set_api(self, api):
        self.api = api

    def generate_input_column_list(self):
        column_list = [
            InputColumnModel(key=CABLE_DESIGN_NAME_KEY, required=True),
            InputColumnModel(key=CABLE_DESIGN_LAYING_KEY, label=LABEL_CABLES_LAYING, required=True),
            InputColumnModel(key=CABLE_DESIGN_VOLTAGE_KEY, label=LABEL_CABLES_VOLTAGE, required=True),
            InputColumnModel(key=CABLE_DESIGN_OWNER_KEY, label=LABEL_CABLES_OWNER, required=True),
            InputColumnModel(key=CABLE_DESIGN_TYPE_KEY, label=LABEL_CABLES_TYPE, required=True),
            InputColumnModel(key=CABLE_DESIGN_SRC_LOCATION_KEY, label=LABEL_CABLES_SRC_LOCATION, required=True),
            InputColumnModel(key=CABLE_DESIGN_SRC_ANS_KEY, label=LABEL_CABLES_SRC_ANSU, required=True),
            InputColumnModel(key=CABLE_DESIGN_SRC_ETPM_KEY, label=LABEL_CABLES_SRC_ETPMC, required=True),
            InputColumnModel(key=CABLE_DESIGN_SRC_ADDRESS_KEY, label=LABEL_CABLES_SRC_ADDRESS, required=True),
            InputColumnModel(key=CABLE_DESIGN_SRC_DESCRIPTION_KEY, label=LABEL_CABLES_SRC_DESCRIPTION, required=True),
            InputColumnModel(key=CABLE_DESIGN_DEST_LOCATION_KEY, label=LABEL_CABLES_DEST_LOCATION, required=True),
            InputColumnModel(key=CABLE_DESIGN_DEST_ANS_KEY, label=LABEL_CABLES_DEST_ANSU, required=True),
            InputColumnModel(key=CABLE_DESIGN_DEST_ETPM_KEY, label=LABEL_CABLES_DEST_ETPMC, required=True),
            InputColumnModel(key=CABLE_DESIGN_DEST_ADDRESS_KEY, label=LABEL_CABLES_DEST_ADDRESS, required=True),
            InputColumnModel(key=CABLE_DESIGN_DEST_DESCRIPTION_KEY, label=LABEL_CABLES_DEST_DESCRIPTION, required=True),
            InputColumnModel(key=CABLE_DESIGN_CABLE_ID_KEY, label=LABEL_CABLES_CABLE_ID),
            InputColumnModel(key=CABLE_DESIGN_END1_DEVICE_NAME_KEY, label=LABEL_CABLES_END1_DEVICE, required=True),
            InputColumnModel(key=CABLE_DESIGN_END1_PORT_NAME_KEY, label=LABEL_CABLES_END1_PORT, required=False),
            InputColumnModel(key=CABLE_DESIGN_END2_DEVICE_NAME_KEY, label=LABEL_CABLES_END2_DEVICE, required=True),
            InputColumnModel(key=CABLE_DESIGN_END2_PORT_NAME_KEY, label=LABEL_CABLES_END2_PORT, required=False),
            InputColumnModel(key=CABLE_DESIGN_IMPORT_ID_KEY, label=LABEL_CABLES_IMPORT_ID, required=True),
            InputColumnModel(key=CABLE_DESIGN_VIA_ROUTE_KEY, label=LABEL_CABLES_FIRST_WAYPOINT, required=False),
            InputColumnModel(key=CABLE_DESIGN_WAYPOINT_ROUTE_KEY, label=LABEL_CABLES_FINAL_WAYPOINT, required=False),
            InputColumnModel(key=CABLE_DESIGN_NOTES_KEY, label=LABEL_CABLES_NOTES, required=False),
        ]
        return column_list

    def generate_output_column_list(self):
        CableDesignOutputObject.ignore_port_columns = self.ignore_port_columns
        return CableDesignOutputObject.get_output_columns()

    def generate_handler_list(self):
        handler_list = [
            UniqueNameHandler(CABLE_DESIGN_IMPORT_ID_KEY, self.info_manager.cable_design_names),
            NamedRangeHandler(self.info_manager, CABLE_DESIGN_LAYING_KEY, "_Laying"),
            NamedRangeHandler(self.info_manager, CABLE_DESIGN_VOLTAGE_KEY, "_Voltage"),
            NamedRangeHandler(self.info_manager, CABLE_DESIGN_OWNER_KEY, "_Owner"),
            CableTypeValidForTechnicalSystemHandler(self.info_manager, CABLE_DESIGN_TYPE_KEY),
            NamedRangeHandler(self.info_manager, CABLE_DESIGN_SRC_LOCATION_KEY, "_Location"),
            ConnectedMenuHandler(self.info_manager, CABLE_DESIGN_SRC_ANS_KEY, CABLE_DESIGN_SRC_LOCATION_KEY),
            ConnectedMenuHandler(self.info_manager, CABLE_DESIGN_SRC_ETPM_KEY, CABLE_DESIGN_SRC_ANS_KEY),
            DeviceAddressHandler(self.info_manager, CABLE_DESIGN_SRC_ADDRESS_KEY, CABLE_DESIGN_SRC_LOCATION_KEY, CABLE_DESIGN_SRC_ETPM_KEY),
            NamedRangeHandler(self.info_manager, CABLE_DESIGN_DEST_LOCATION_KEY, "_Location"),
            ConnectedMenuHandler(self.info_manager, CABLE_DESIGN_DEST_ANS_KEY, CABLE_DESIGN_DEST_LOCATION_KEY),
            ConnectedMenuHandler(self.info_manager, CABLE_DESIGN_DEST_ETPM_KEY, CABLE_DESIGN_DEST_ANS_KEY),
            DeviceAddressHandler(self.info_manager, CABLE_DESIGN_DEST_ADDRESS_KEY, CABLE_DESIGN_DEST_LOCATION_KEY, CABLE_DESIGN_DEST_ETPM_KEY),
        ]

        if not self.validate_only:

            cable_id_index = self.get_input_column_index_for_key(CABLE_DESIGN_CABLE_ID_KEY)
            import_id_index = self.get_input_column_index_for_key(CABLE_DESIGN_IMPORT_ID_KEY)
            handler_list.append(CableDesignExistenceHandler(CABLE_DESIGN_IMPORT_ID_KEY,
                                                            self.info_manager,
                                                            cable_id_index,
                                                            import_id_index))

            end1_device_name_index = self.get_input_column_index_for_key(CABLE_DESIGN_END1_DEVICE_NAME_KEY)
            src_etpm_index = self.get_input_column_index_for_key(CABLE_DESIGN_SRC_ETPM_KEY)
            handler_list.append(EndpointHandler(CABLE_DESIGN_END1_DEVICE_NAME_KEY,
                                                CABLE_DESIGN_SRC_ETPM_KEY,
                                                self.get_md_root(),
                                                self.info_manager,
                                                self.api,
                                                end1_device_name_index,
                                                src_etpm_index,
                                                "source endpoints"))

            handler_list.append(DevicePortHandler(CABLE_DESIGN_END1_PORT_NAME_KEY,
                                                  CABLE_DESIGN_END1_DEVICE_NAME_KEY,
                                                  self.info_manager,
                                                  self.ignore_port_columns,
                                                  self.info_manager.from_port_values))

            end2_device_name_index = self.get_input_column_index_for_key(CABLE_DESIGN_END2_DEVICE_NAME_KEY)
            dest_etpm_index = self.get_input_column_index_for_key(CABLE_DESIGN_DEST_ETPM_KEY)
            handler_list.append(EndpointHandler(CABLE_DESIGN_END2_DEVICE_NAME_KEY,
                                                CABLE_DESIGN_DEST_ETPM_KEY,
                                                self.get_md_root(),
                                                self.info_manager,
                                                self.api,
                                                end2_device_name_index,
                                                dest_etpm_index,
                                                "destination endpoints"))

            handler_list.append(DevicePortHandler(CABLE_DESIGN_END2_PORT_NAME_KEY,
                                                  CABLE_DESIGN_END2_DEVICE_NAME_KEY,
                                                  self.info_manager,
                                                  self.ignore_port_columns,
                                                  self.info_manager.to_port_values))

        return handler_list

    def get_md_root(self):
        return self.md_root

    @classmethod
    def blank_row_columns_allowed_dict(cls):
        return {CABLE_DESIGN_NAME_KEY: "[] | []",
                CABLE_DESIGN_IMPORT_ID_KEY: None,
                CABLE_DESIGN_NOTES_KEY: None}

    def handle_valid_row(self, input_dict):

        cable_catalog_info = self.info_manager.get_cable_type_info(input_dict[CABLE_DESIGN_TYPE_KEY])

        # add output object for cable design comparison tab in output workbook
        cable_design_name = name = CableDesignOutputObject.get_name_cls(input_dict)
        if cable_design_name in self.info_manager.existing_cable_designs:
            cable_catalog_info = self.info_manager.get_cable_type_info(input_dict[CABLE_DESIGN_TYPE_KEY])
            cable_design_id = self.info_manager.id_manager_cable_design.get_id_for_name(cable_design_name)
            self.info_manager.get_output_objects_cable_design_compare().append(
                CableDesignOutputObject(
                    helper=self,
                    input_dict=input_dict,
                    cable_catalog_info=cable_catalog_info,
                    ignore_port_columns=self.ignore_port_columns,
                    existing_item_id=cable_design_id))

        # add output object for cable inventory import tab in output workbook
        cable_type_name = cable_catalog_info.name
        self.info_manager.get_output_objects_cable_inventory().append(
            CableInventoryOutputObject(helper=self,
                                       input_dict=None,
                                       tag_name=cable_design_name,
                                       cable_type_name=cable_type_name)
        )

        # add output object for cable design import tab in output workbook
        logging.debug("adding output object for: %s" % input_dict[CABLE_DESIGN_NAME_KEY])
        self.output_objects.append(CableDesignOutputObject(helper=self,
                                                           input_dict=input_dict,
                                                           cable_catalog_info=cable_catalog_info,
                                                           ignore_port_columns=self.ignore_port_columns))

    def get_summary_messages_custom(self):

        messages = []

        messages.append("Cable designs that already exist in CDB: %d" % len(self.info_manager.existing_cable_designs))
        messages.append("Cable designs that need to be added to CDB: %d" % len(self.info_manager.new_cable_designs))

        num_port_values = len(self.info_manager.from_port_values) + len(self.info_manager.to_port_values)
        if num_port_values > 0:
            messages.append("WARNING: ignored %d non-empty values in from/to device port columns (ignorePortColumns config resource set to true)" % num_port_values)

        return messages

    def get_summary_sheet_columns(self):

        summary_column_names = []
        summary_column_values = []

        if len(self.info_manager.existing_cable_designs) > 0:
            summary_column_names.append("existing cable designs")
            summary_column_values.append(self.info_manager.existing_cable_designs)

        if len(self.info_manager.new_cable_designs) > 0:
            summary_column_names.append("new cable designs")
            summary_column_values.append(self.info_manager.new_cable_designs)

        ignored_ports = self.info_manager.from_port_values + self.info_manager.to_port_values
        if len(ignored_ports) > 0:
            summary_column_names.append("ignored from/to ports")
            summary_column_values.append(ignored_ports)

        return summary_column_names, summary_column_values

    def write_helper_sheets(self, output_book):

        # create list of catalog port output objects and use to generate tab for catalog ports in output workbook
        catalog_port_output_objects = []
        for catalog_name, catalog_ports in self.info_manager.catalog_ports.items():
            for port_name in catalog_ports:
                if port_name is not None and port_name != "":
                    output_object = CatalogPortOutputObject(catalog_name, port_name)
                    catalog_port_output_objects.append(output_object)
        self.write_sheet(output_book, "Catalog Port Import", CatalogPortOutputObject.get_output_columns(), catalog_port_output_objects)

        self.write_sheet(output_book, "Cable Inventory Item Import", CableInventoryOutputObject.get_output_columns(), self.info_manager.get_output_objects_cable_inventory())
        self.write_sheet(output_book, "Cable Design Item Compare", self.output_column_list(), self.info_manager.get_output_objects_cable_design_compare())
        self.write_sheet(output_book, "Cable Design Item Import", self.output_column_list(), self.output_objects)

    def get_error_messages_custom(self):

        messages = []

        if len(self.info_manager.missing_endpoints) > 0:
            messages.append("Endpoint device names not defined in CDB: %d" % len(self.info_manager.missing_endpoints))

        if len(self.info_manager.nonunique_endpoints) > 0:
            messages.append("Multiple CDB devices with matching name: %d" % len(self.info_manager.nonunique_endpoints))

        return messages

    def get_error_sheet_columns(self):

        error_column_names = []
        error_column_values = []

        if len(self.info_manager.missing_endpoints) > 0:
            error_column_names.append("missing endpoints")
            error_column_values.append(sorted(self.info_manager.missing_endpoints))

        if len(self.info_manager.nonunique_endpoints) > 0:
            error_column_names.append("non-unique endpoints")
            error_column_values.append(sorted(self.info_manager.nonunique_endpoints))

        return error_column_names, error_column_values

    # Returns processing summary message.
    def get_processing_summary(self):

        processing_summary = ""

        if len(self.info_manager.existing_cable_designs) > 0:
            processing_summary = processing_summary + "DETAILS: number of cable designs that already exist in CDB (included in output file): %d" % (len(self.info_manager.existing_cable_designs))

        num_port_values = len(self.info_manager.from_port_values) + len(self.info_manager.to_port_values)
        if num_port_values > 0:
            if processing_summary != "":
                processing_summary = processing_summary + "\n"
            processing_summary = processing_summary + "WARNING: ignored %d non-empty values in from/to device port columns (ignorePortColumns config resource set to true)" % num_port_values

        return processing_summary


range_parts = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')
def xl_cell_to_rowcol(cell_str):
    """
    NOTE: I borrowed this from the xlsxwriter library since it was useful for parsing Excel $G$1 notation.
    Convert a cell reference in A1 notation to a zero indexed row and column.

    Args:
       cell_str:  A1 style string.

    Returns:
        row, col: Zero indexed cell row and column indices.

    """
    if not cell_str:
        return 0, 0

    match = range_parts.match(cell_str)
    col_str = match.group(2)
    row_str = match.group(4)

    # Convert base26 column string to number.
    expn = 0
    col = 0
    for char in reversed(col_str):
        col += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1

    row = int(row_str)

    return row, col


def get_row_height_wrapped_message(message):
    return ((len(message) // DEFAULT_COLUMN_WIDTH) + 2) * DEFAULT_FONT_HEIGHT


def fatal_error(error_msg):
    print()
    print("ERROR ====================")
    print()
    print(error_msg)
    sys.exit(0)


def get_config_resource(config, section, key, is_required, print_value=True, print_mask=None):
    value = None
    if section not in config:
        fatal_error("Invalid config section: %s, exiting" % section)
    if key not in config[section]:
        if is_required:
            fatal_error("Config key: %s not found in section: %s, exiting" % (key, section))
    else:
        value = config[section][key]
        if is_required and len(value) == 0:
            fatal_error("value not provided for required option '[%s] %s', exiting" % (section, key))
    if print_value:
        print_obj = value
        if value is not None:
            if print_mask is not None:
                if len(value) > 0:
                    print_obj = print_mask
        print("[%s] %s: %s" % (section, key, print_obj))
    return value


def get_config_resource_boolean(config, section, key, is_required, print_value=True, print_mask=None):
    config_value = get_config_resource(config, section, key, is_required, print_value=False)
    boolean_value = False
    if config_value in ("True", "TRUE", "true", "Yes", "YES", "yes", "On", "ON", "on", "1"):
        boolean_value = True
    print("[%s] %s: %s" % (section, key, boolean_value))
    return boolean_value


def get_config_resource_int(config, section, key, is_required, print_value=True, print_mask=None):
    config_value = get_config_resource(config, section, key, is_required, print_value=False)
    if config_value is None:
        return 0
    else:
        return int(config_value)


def main():

    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument("--configDir", help="Directory containing script config files.", required=True)
    parser.add_argument("--kabelWorkbookId", help="Symbolic identifier for kabel workbook, must have corresponding '.conf' file in configDir", required=True)
    parser.add_argument("--deploymentName", help="Name to use for looking up URL/user/password in deploymentInfoFile")
    parser.add_argument("--cdbUrl", help="CDB system URL")
    parser.add_argument("--cdbUser", help="CDB User ID for API login")
    parser.add_argument("--cdbPassword", help="CDB User password for API login")
    args = parser.parse_args()

    print()
    print("COMMAND LINE ARGS ====================")
    print()
    print("configDir: %s" % args.configDir)
    print("kabelWorkbookId: %s" % args.kabelWorkbookId)
    print("deploymentName: %s" % args.deploymentName)
    print("cdbUrl: %s" % args.cdbUrl)
    print("cdbUser: %s" % args.cdbUser)
    if args.cdbPassword is not None and len(args.cdbPassword) > 0:
        password_string = "********"
    else:
        password_string = args.cdbPassword
    print("cdbPassword: %s" % password_string)

    #
    # determine config file names and paths and test
    #

    option_config_dir = args.configDir
    if not os.path.isdir(option_config_dir):
        fatal_error("Specified configDir: %s does not exist, exiting" % option_config_dir)

    file_config_preimport = option_config_dir + "/preimport.conf"
    if not os.path.isfile(file_config_preimport):
        fatal_error("'preimport.conf' file not found in configDir: %s', exiting" % option_config_dir)

    file_config_deployment_info = option_config_dir + "/cdb-deployment-info.conf"

    option_workbook_id = args.kabelWorkbookId
    file_config_workbook = option_config_dir + "/" + option_workbook_id + ".conf"
    if not os.path.isfile(file_config_workbook):
        fatal_error("'%s.conf' file not found in configDir: %s', exiting" % (option_workbook_id, option_config_dir))

    #
    # process options
    #

    # read config files
    config_preimport = configparser.ConfigParser()
    config_preimport.read(file_config_preimport)
    config_workbook = configparser.ConfigParser()
    config_workbook.read(file_config_workbook)

    print()
    print("preimport.conf OPTIONS ====================")
    print()

    # process inputDir option
    option_input_dir = get_config_resource(config_preimport, 'DEFAULT', 'inputDir', True)
    if not os.path.isdir(option_input_dir):
        fatal_error("'[%s] inputDir' directory: %s does not exist, exiting" % ('DEFAULT', option_input_dir))

    # process outputDir option
    option_output_dir = get_config_resource(config_preimport, 'DEFAULT', 'outputDir', True)
    if not os.path.isdir(option_output_dir):
        fatal_error("'[%s] outputDir' directory: %s does not exist, exiting" % ('DEFAULT', option_output_dir))

    print()
    print("%s.conf OPTIONS ====================" % option_workbook_id)
    print()

    # process inputExcelFile option
    option_input_file = get_config_resource(config_workbook, 'DEFAULT', 'inputExcelFile', True)
    file_input = option_input_dir + "/" + option_input_file
    if not os.path.isfile(file_input):
        fatal_error("'[%s] inputExcelFile' file: %s does not exist in directory: %s, exiting" % ('DEFAULT', option_input_file, option_input_dir))

    # process validateOnly option
    validate_only = get_config_resource_boolean(config_workbook, 'DEFAULT', 'validateOnly', False)

    #
    # Generate output file paths.
    #

    # output excel file
    file_output = "%s/%s.xlsx" % (option_output_dir, option_workbook_id)

    # log file
    file_log = "%s/%s.log" % (option_output_dir, option_workbook_id)

    #
    # determine whether to use args or config for url/user/password
    #

    # get cdb url, user, password from config, if specified
    option_cdb_url = None
    option_cdb_user = None
    option_cdb_password = None

    if len(args.deploymentName) > 0:
        if len(file_config_deployment_info) == 0:
            # must have deployment info file
            fatal_error("[INPUTS] deploymentInfoFile not specified but required to look up deployment name: %s, exiting" % args.deploymentName)
        else:
            if not os.path.isfile(file_config_deployment_info):
                fatal_error("'[INPUTS] deploymentInfoFile' file: %s does not exist, exiting" % file_config_deployment_info)
            else:
                print()
                print("DEPLOYMENT INFO CONFIG ====================")
                print()
                deployment_config = configparser.ConfigParser()
                deployment_config.read(file_config_deployment_info)
                if args.deploymentName not in deployment_config:
                    fatal_error("specified deploymentName: %s not found in deploymentInfoFile: %s, exiting" % (args.deploymentName, file_config_deployment_info))
                option_cdb_url = get_config_resource(deployment_config, args.deploymentName, 'cdbUrl', False)
                option_cdb_user = get_config_resource(deployment_config, args.deploymentName, 'cdbUser', False)
                option_cdb_password = get_config_resource(deployment_config, args.deploymentName, 'cdbPassword', False, True, '********')

    if args.cdbUrl is not None:
        cdb_url = args.cdbUrl
    else:
        if option_cdb_url is None:
            fatal_error("cdbUser must be specified on command line or via deployment info file, exiting")
        else:
            cdb_url = option_cdb_url

    if args.cdbUser is not None:
        cdb_user = args.cdbUser
    else:
        if option_cdb_user is None:
            fatal_error("cdbUser must be specified on command line or via deployment info file, exiting")
        else:
            cdb_user = option_cdb_user

    if args.cdbPassword is not None:
        cdb_password = args.cdbPassword
    else:
        if option_cdb_password is None:
            fatal_error("cdbUser must be specified on command line or via deployment info file, exiting")
        else:
            cdb_password = option_cdb_password

    print()
    print("CDB URL/USER/PASSWORD SETTINGS (COMMAND LINE TAKES PRECEDENCE OVER CONFIG)====================")
    print()
    print("cdbUrl: %s" % cdb_url)
    print("cdbUser: %s" % cdb_user)
    print("cdbPassword: %s" % cdb_password)
    if cdb_password is not None and len(cdb_password) > 0:
        password_string = "********"
    else:
        password_string = cdb_password
    print("cdbPassword: %s" % password_string)

    # configure logging
    logging.basicConfig(filename=file_log, filemode='w', level=logging.DEBUG, format='%(levelname)s - %(message)s')

    #
    # connect to CDB
    #

    print()
    print("CONNECTING TO CDB ====================")
    print()
    print("connecting to %s as user %s (make sure CDB is running and VPN connected if needed)" % (cdb_url, cdb_user))

    # open connection to CDB
    api = CdbApiFactory(cdb_url)
    try:
        api.authenticateUser(cdb_user, cdb_password)
        api.testAuthenticated()
    except (ApiException, urllib3.exceptions.MaxRetryError) as exc:
        fatal_error("CDB login failed user: %s message: %s, exiting" % (cdb_user, exc))

    # open input spreadsheet
    input_book = openpyxl.load_workbook(filename=file_input, data_only=True)

    # create output spreadsheet
    output_book = xlsxwriter.Workbook(file_output)

    # create information manager for retrieving item data via cdb api and managing cdb and parsed item data
    info_manager = ItemInfoManager(api)
    info_manager.initialize(input_book)

    # create sheet helpers
    helpers = [CableSpecsSheetHelper(), CablesSheetHelper()]

    for helper in helpers:
        print()
        print("%s OPTIONS ====================" % helper.sheet_name().upper())
        print()
        helper.post_create(config_preimport, config_workbook, info_manager, api)
        input_valid = helper.process_input_book(input_book)
        if not input_valid:
            helper.write_error_sheets(output_book)
            helper.write_validation_sheet(output_book)
            break

    if input_valid:
        for helper in helpers:
            if not validate_only:
                helper.write_workbook_sheets(output_book)
            else:
                helper.write_validate_only_sheets(output_book)

    # clean up
    output_book.close()

    # close CDB connection
    try:
        api.logOutUser()
    except ApiException:
        logging.error("CDB logout failed")

    print()
    print("OUTPUT FILES ====================")
    print()
    print("output workbook: %s" % file_output)
    print("log file: %s" % file_log)


if __name__ == '__main__':
    main()